import os, uuid, json, re, logging, asyncio
from datetime import datetime, timezone, timedelta
from logging.handlers import TimedRotatingFileHandler

from fastapi import FastAPI, UploadFile, Form, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import FileResponse, Response
import httpx, secrets
from openai import AsyncOpenAI
from dotenv import load_dotenv
from db import init_db, get_conn
from zoom_client import ZoomClient
from zoom_webhook import verify_signature, build_url_validation_response

load_dotenv()
init_db()

# ── ロギング設定 ──────────────────────────────────
LOG_DIR  = '/var/log/jizo-api'
LOG_FILE = os.path.join(LOG_DIR, 'app.log')
os.makedirs(LOG_DIR, exist_ok=True)

_fmt = logging.Formatter('%(asctime)s %(levelname)s: %(message)s',
                          datefmt='%Y-%m-%d %H:%M:%S')
# 12時間ごとにローテーション・2世代保持（=最大24時間分）
_fh = TimedRotatingFileHandler(LOG_FILE, when='H', interval=12, backupCount=2, encoding='utf-8')
_fh.setFormatter(_fmt)
_fh.setLevel(logging.DEBUG)

logger = logging.getLogger('jizo-api')
logger.setLevel(logging.DEBUG)
logger.addHandler(_fh)
logger.propagate = False  # uvicornのrootロガーに伝播させない

API_KEY     = os.environ['ASSEMBLYAI_API_KEY']
AAI_HDR     = {'authorization': API_KEY}
AAI_BASE    = 'https://api.assemblyai.com'
GEMINI_KEY    = os.environ.get('GEMINI_API_KEY', '')
GEMINI_BASE   = 'https://generativelanguage.googleapis.com/v1beta/openai'
OPENAI_KEY    = os.environ.get('OPENAI_API_KEY', '')
DEEPGRAM_KEY  = os.environ.get('DEEPGRAM_API_KEY', '')
DEEPGRAM_BASE = 'https://api.deepgram.com/v1'
ANTHROPIC_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
ANTHROPIC_BASE = 'https://api.anthropic.com/v1'

# ── Zoom 連携（v1） ──────────────────────────────
ZOOM_CLIENT_ID      = os.environ.get('ZOOM_CLIENT_ID', '')
ZOOM_CLIENT_SECRET  = os.environ.get('ZOOM_CLIENT_SECRET', '')
ZOOM_ACCOUNT_ID     = os.environ.get('ZOOM_ACCOUNT_ID', '')
ZOOM_WEBHOOK_SECRET = os.environ.get('ZOOM_WEBHOOK_SECRET', '')
ZOOM_AUDIO_DIR      = '/var/jizo/audio/zoom'
os.makedirs(ZOOM_AUDIO_DIR, exist_ok=True)

# 環境変数が未設定なら None（Webhook受信時に明示エラー）
zoom_client = ZoomClient(ZOOM_CLIENT_ID, ZOOM_CLIENT_SECRET, ZOOM_ACCOUNT_ID) \
              if ZOOM_CLIENT_ID and ZOOM_CLIENT_SECRET and ZOOM_ACCOUNT_ID else None

# ── 使用量上限（無料枠ベース・デフォルト）─────────
USAGE_LIMITS = {
    'whisper': {
        'daily_minutes':   30,
        'monthly_minutes': 100,
        'rate_usd_per_min': 0.006,
        'free_tier_note':  '事前チャージ（最低$5）',
    },
    'deepgram': {
        'daily_minutes':   60,
        'monthly_minutes': 750,
        'rate_usd_per_min': 0.0043,
        'free_tier_note':  '$200無料クレジット（約12時間/750分）',
    },
    'gemini-flash': {
        'daily_requests': 1000,
        'daily_tokens':   800_000,
        'monthly_requests': 30_000,
        'rate_usd_per_1m_in':  0.30,
        'rate_usd_per_1m_out': 2.50,
        'free_tier_note': '無料枠: 1500req/日・15req/分',
    },
    'gemini-pro': {
        'daily_requests': 80,
        'daily_tokens':   200_000,
        'monthly_requests': 2_400,
        'rate_usd_per_1m_in':  1.25,
        'rate_usd_per_1m_out': 10.00,
        'free_tier_note': '無料枠: 100req/日・5req/分',
    },
    'claude-haiku': {
        'daily_requests': 500,
        'daily_tokens':   1_000_000,
        'monthly_requests': 15_000,
        'rate_usd_per_1m_in':  1.00,
        'rate_usd_per_1m_out': 5.00,
        'free_tier_note': '前払いクレジット（最小$5）',
    },
}

def _service_key(engine_or_model: str) -> str:
    """エンジン名/モデル名から使用量集計キーを返す"""
    e = (engine_or_model or '').lower()
    if e == 'whisper' or 'whisper' in e: return 'whisper'
    if e == 'deepgram' or 'nova' in e:   return 'deepgram'
    if 'haiku' in e and 'claude' in e:   return 'claude-haiku'
    if 'flash' in e: return 'gemini-flash'
    if 'pro'   in e: return 'gemini-pro'
    return e or 'unknown'


def _record_usage(service: str, units: float, unit_type: str,
                  project_id: str = None, endpoint: str = None,
                  cost_usd: float = 0.0):
    """使用量を api_usage テーブルに記録"""
    try:
        c = get_conn()
        c.execute(
            'INSERT INTO api_usage(service,units,unit_type,project_id,endpoint,cost_usd,created_at) '
            'VALUES(?,?,?,?,?,?,?)',
            (service, units, unit_type, project_id, endpoint, cost_usd, now_iso())
        )
        c.commit()
        c.close()
        logger.info(f'usage: {service}={units}{unit_type} cost=${cost_usd:.4f}')
    except Exception as e:
        logger.warning(f'usage record failed: {e}')


def _get_usage(service: str, period: str = 'day') -> dict:
    """指定サービス・期間（day/month）の使用量集計"""
    from datetime import timedelta
    now = datetime.now(timezone.utc)
    if period == 'day':
        since = (now - timedelta(days=1)).isoformat()
    else:
        since = (now - timedelta(days=30)).isoformat()
    c = get_conn()
    rows = c.execute(
        'SELECT unit_type, SUM(units) AS u, SUM(cost_usd) AS cost, COUNT(*) AS cnt '
        'FROM api_usage WHERE service=? AND created_at>=? GROUP BY unit_type',
        (service, since)
    ).fetchall()
    c.close()
    result = {'requests': 0, 'minutes': 0.0, 'tokens': 0, 'cost_usd': 0.0}
    for r in rows:
        d = dict(r)
        if d['unit_type'] == 'minutes':  result['minutes'] = d['u']
        elif d['unit_type'] == 'tokens': result['tokens']  = int(d['u'])
        result['cost_usd'] += d['cost'] or 0
        result['requests'] += d['cnt']
    return result


def _check_limit(service: str, expected_units: float = 0, expected_unit_type: str = 'minutes'):
    """上限超過チェック。超えていたら HTTPException(429) raise"""
    limits = USAGE_LIMITS.get(service, {})
    if not limits: return  # 未定義サービスはスルー

    daily   = _get_usage(service, 'day')
    monthly = _get_usage(service, 'month')

    if expected_unit_type == 'minutes':
        if 'daily_minutes' in limits and daily['minutes'] + expected_units > limits['daily_minutes']:
            raise HTTPException(429,
                f'{service} 日次上限超過：{daily["minutes"]:.1f}分 / {limits["daily_minutes"]}分'
                f'（予定追加: {expected_units:.1f}分）')
        if 'monthly_minutes' in limits and monthly['minutes'] + expected_units > limits['monthly_minutes']:
            raise HTTPException(429,
                f'{service} 月次上限超過：{monthly["minutes"]:.1f}分 / {limits["monthly_minutes"]}分')
    elif expected_unit_type == 'requests':
        if 'daily_requests' in limits and daily['requests'] + 1 > limits['daily_requests']:
            raise HTTPException(429,
                f'{service} 日次リクエスト上限超過：{daily["requests"]} / {limits["daily_requests"]}')
AUDIO_DIR   = os.environ.get('AUDIO_DIR', '/var/jizo/audio')
ADMIN_USER  = os.environ.get('ADMIN_USER', '')
ADMIN_PASS  = os.environ.get('ADMIN_PASS', '')

os.makedirs(AUDIO_DIR, exist_ok=True)

app = FastAPI()
# auto_error=False: Authorization ヘッダが無くても 401 を自動送出せず（=ブラウザの
# ネイティブBasic認証ダイアログを出さず）、require_admin 側で WWW-Authenticate 無しの
# 401 を返す。認証は管理画面の独自ログインフォームで行う。
security = HTTPBasic(auto_error=False)

app.add_middleware(
    CORSMiddleware,
    allow_origins=['https://jizo-dev.com'],
    allow_methods=['*'],
    allow_headers=['*'],
)

def now_iso():
    return datetime.now(timezone.utc).isoformat()

# 音声アップロードの content_type / 拡張子 → 保存拡張子
_MIME_EXT = {
    'audio/webm': 'webm', 'audio/wav': 'wav', 'audio/x-wav': 'wav',
    'audio/mpeg': 'mp3', 'audio/mp3': 'mp3',
    'audio/mp4': 'mp4', 'audio/m4a': 'm4a', 'audio/x-m4a': 'm4a',
    'audio/ogg': 'ogg', 'audio/flac': 'flac', 'audio/x-flac': 'flac',
}
def _ext_for_upload(upload):
    mt = (upload.content_type or '').lower()
    if mt in _MIME_EXT:
        return _MIME_EXT[mt]
    fn = (upload.filename or '').lower()
    for ext in ('webm', 'wav', 'mp3', 'm4a', 'mp4', 'ogg', 'flac'):
        if fn.endswith('.' + ext):
            return ext
    return 'webm'

def require_admin(creds: HTTPBasicCredentials | None = Depends(security)):
    # 認証情報が無い／管理者資格情報が未設定なら 401（WWW-Authenticate は付けない＝
    # ブラウザのネイティブBasicダイアログを出さない。空設定でのバイパスも防ぐ）。
    if creds is None or not ADMIN_USER or not ADMIN_PASS:
        raise HTTPException(401, 'Unauthorized')
    ok = (
        secrets.compare_digest(creds.username.encode(), ADMIN_USER.encode()) and
        secrets.compare_digest(creds.password.encode(), ADMIN_PASS.encode())
    )
    if not ok:
        raise HTTPException(401, 'Unauthorized')
    return creds.username


# ── ヘルス ──────────────────────────────────────────
@app.get('/api/health')
async def health():
    logger.info('health check')
    return {'ok': True}


# ── ログ閲覧（管理画面・Claude Code用） ──────────────
@app.get('/api/logs')
async def get_logs(lines: int = 200, _: str = Depends(require_admin)):
    try:
        with open(LOG_FILE, encoding='utf-8', errors='replace') as f:
            all_lines = f.readlines()
        tail = all_lines[-lines:] if len(all_lines) > lines else all_lines
        return {'lines': len(tail), 'log': ''.join(tail)}
    except FileNotFoundError:
        return {'lines': 0, 'log': '(ログファイルなし)'}


# ── プロジェクト作成（モバイルからのアップロード） ──────
@app.post('/api/projects')
async def create_project(
    meta: str = Form(...),
    audio: list[UploadFile] = [],
):
    """
    multipart/form-data:
      meta  : JSON文字列 { name, participants:[str], segments:[{text,speakerIdx,ts,highlight}] }
      audio : 音声ファイル（0個以上、停止/再開ごとの複数チャンク）
    """
    try:
        m = json.loads(meta)
    except Exception:
        raise HTTPException(400, 'meta must be valid JSON')

    project_id   = str(uuid.uuid4())
    name           = m.get('name', '')
    participants   = m.get('participants', [])
    segments       = m.get('segments', [])
    speaker_events = m.get('speaker_events', []) or []

    conn = get_conn()
    try:
        conn.execute(
            'INSERT INTO projects(id,name,created_at,status) VALUES(?,?,?,?)',
            (project_id, name, now_iso(), 'uploaded')
        )
        for idx, pname in enumerate(participants):
            conn.execute(
                'INSERT INTO participants(project_id,idx,name) VALUES(?,?,?)',
                (project_id, idx, pname or '')
            )
        for i, seg in enumerate(segments):
            conn.execute(
                'INSERT INTO ref_segments(project_id,seq,text,speaker_idx,ts,highlight) VALUES(?,?,?,?,?,?)',
                (project_id, i, seg.get('text', ''), seg.get('speakerIdx'),
                 seg.get('ts', ''), int(bool(seg.get('highlight', False))))
            )
        # 話者ボタン押下イベント（LLM突合の権威データ）
        for i, ev in enumerate(speaker_events):
            conn.execute(
                'INSERT INTO speaker_events(project_id,seq,ts,ms,speaker_idx,speaker_name) VALUES(?,?,?,?,?,?)',
                (project_id, i,
                 ev.get('ts', ''),
                 int(ev.get('ms') or 0),
                 ev.get('speaker_idx'),
                 ev.get('speaker_name', ''))
            )
        _MIME_EXT = {
            'audio/webm': 'webm', 'audio/wav': 'wav', 'audio/x-wav': 'wav',
            'audio/mpeg': 'mp3', 'audio/mp3': 'mp3',
            'audio/mp4': 'mp4', 'audio/m4a': 'm4a', 'audio/x-m4a': 'm4a',
            'audio/ogg': 'ogg', 'audio/flac': 'flac', 'audio/x-flac': 'flac',
        }
        def _ext_for(upload):
            mt = (upload.content_type or '').lower()
            if mt in _MIME_EXT: return _MIME_EXT[mt]
            fn = (upload.filename or '').lower()
            for ext in ('webm','wav','mp3','m4a','mp4','ogg','flac'):
                if fn.endswith('.' + ext): return ext
            return 'webm'

        saved_seq = 0
        for f in audio:
            data = await f.read()
            if not data:
                logger.warning(f'create_project: empty audio chunk skipped (project={project_id}, filename={f.filename})')
                continue
            ext = _ext_for(f)
            path = os.path.join(AUDIO_DIR, f'{project_id}_{saved_seq}.{ext}')
            with open(path, 'wb') as fp:
                fp.write(data)
            conn.execute(
                'INSERT INTO recordings(project_id,seq,audio_path,mime,duration,created_at) VALUES(?,?,?,?,?,?)',
                (project_id, saved_seq, path, f.content_type or f'audio/{ext}', '', now_iso())
            )
            saved_seq += 1
        conn.commit()
    finally:
        conn.close()

    return {'project_id': project_id, 'recordings': len(audio)}


# ── 録音チャンク追記（モバイルの15分自動分割アップロード用・無認証） ──
@app.post('/api/projects/{project_id}/chunks')
async def append_chunk(
    project_id: str,
    meta: str = Form('{}'),
    audio: list[UploadFile] = [],
):
    """
    既存プロジェクト（同一MTG）へ録音チャンクを追記する。
    モバイルが録音を止めずに 15 分ごとに送信し、同じ MTG に繋げるための無認証 API。
    create_project と同じ信頼境界（モバイルは Basic 資格情報を持たない）。

    multipart/form-data:
      meta  : JSON文字列 { segments:[{text,speakerIdx,ts,highlight}], speaker_events:[...] }
              ※前回送信以降の「差分」のみをクライアントが送る
      audio : 音声ファイル（このチャンク分・通常1個）

    recordings / ref_segments / speaker_events の seq は、それぞれ
    既存の MAX(seq)+1 から連番で追記する（各テーブル独立系列・時系列は ORDER BY seq で担保）。
    participants は増殖させない（初回作成時のみ）。
    """
    try:
        m = json.loads(meta)
    except Exception:
        raise HTTPException(400, 'meta must be valid JSON')
    segments       = m.get('segments', []) or []
    speaker_events = m.get('speaker_events', []) or []

    conn = get_conn()
    try:
        proj = conn.execute('SELECT id FROM projects WHERE id=?', (project_id,)).fetchone()
        if not proj:
            raise HTTPException(404, 'project not found')

        # 各テーブルの seq 継続採番（独立系列）
        rec_base = conn.execute(
            'SELECT COALESCE(MAX(seq),-1) AS m FROM recordings WHERE project_id=?', (project_id,)
        ).fetchone()['m'] + 1
        seg_base = conn.execute(
            'SELECT COALESCE(MAX(seq),-1) AS m FROM ref_segments WHERE project_id=?', (project_id,)
        ).fetchone()['m'] + 1
        ev_base = conn.execute(
            'SELECT COALESCE(MAX(seq),-1) AS m FROM speaker_events WHERE project_id=?', (project_id,)
        ).fetchone()['m'] + 1

        seq = rec_base
        for f in audio:
            data = await f.read()
            if not data:
                logger.warning(f'append_chunk: empty audio chunk skipped (project={project_id}, filename={f.filename})')
                continue
            ext = _ext_for_upload(f)
            path = os.path.join(AUDIO_DIR, f'{project_id}_{seq}.{ext}')
            with open(path, 'wb') as fp:
                fp.write(data)
            conn.execute(
                'INSERT INTO recordings(project_id,seq,audio_path,mime,duration,created_at) VALUES(?,?,?,?,?,?)',
                (project_id, seq, path, f.content_type or f'audio/{ext}', '', now_iso())
            )
            seq += 1

        for i, seg in enumerate(segments):
            conn.execute(
                'INSERT INTO ref_segments(project_id,seq,text,speaker_idx,ts,highlight) VALUES(?,?,?,?,?,?)',
                (project_id, seg_base + i, seg.get('text', ''), seg.get('speakerIdx'),
                 seg.get('ts', ''), int(bool(seg.get('highlight', False))))
            )
        for i, ev in enumerate(speaker_events):
            conn.execute(
                'INSERT INTO speaker_events(project_id,seq,ts,ms,speaker_idx,speaker_name) VALUES(?,?,?,?,?,?)',
                (project_id, ev_base + i, ev.get('ts', ''), int(ev.get('ms') or 0),
                 ev.get('speaker_idx'), ev.get('speaker_name', ''))
            )
        conn.commit()
    finally:
        conn.close()

    return {'project_id': project_id, 'next_seq': seq, 'recordings_added': seq - rec_base}


# ── keyterms生成（参加者名 + カタカナ固有名詞） ─────────
def _build_keyterms(project_id: str, conn) -> list:
    names = [
        r['name'] for r in conn.execute(
            'SELECT name FROM participants WHERE project_id=? AND name != ""', (project_id,)
        ).fetchall()
    ]
    texts = ' '.join(
        r['text'] for r in conn.execute(
            'SELECT text FROM ref_segments WHERE project_id=?', (project_id,)
        ).fetchall()
    )
    # カタカナ3文字以上を固有名詞候補として抽出
    katakana = re.findall(r'[ァ-ヶー]{3,}', texts)
    terms = list({t for t in (names + katakana) if t})[:100]
    return terms


# ── デフォルトRun設定取得（管理画面用）──────────────
@app.get('/api/run-defaults')
async def get_run_defaults(_: str = Depends(require_admin)):
    return DEFAULT_SETTINGS


# ── 設定画面用：APIキー状態・上限・使用量 ────────────
def _mask_key(k: str) -> str:
    if not k: return ''
    if len(k) <= 12: return '*' * len(k)
    return k[:7] + '...' + k[-4:]

# ── Deepgram公式：プロジェクトID・使用量・残高 ──────
_dg_project_id = None

async def _get_dg_project_id():
    global _dg_project_id
    if _dg_project_id:
        return _dg_project_id
    if not DEEPGRAM_KEY:
        raise HTTPException(503, 'Deepgram key not configured')
    async with httpx.AsyncClient(timeout=15) as client:
        r = await client.get(
            f'{DEEPGRAM_BASE}/projects',
            headers={'Authorization': f'Token {DEEPGRAM_KEY}'},
        )
        if r.status_code != 200:
            raise HTTPException(502, f'Deepgram projects: {r.text[:300]}')
        projs = r.json().get('projects') or []
        if not projs:
            raise HTTPException(500, 'No Deepgram projects found')
        _dg_project_id = projs[0]['project_id']
        return _dg_project_id


@app.get('/api/deepgram-stats')
async def get_deepgram_stats(days: int = 30, _: str = Depends(require_admin)):
    """Deepgram公式集計（使用量＋残高）を取得"""
    pid = await _get_dg_project_id()
    now   = datetime.now(timezone.utc)
    start = (now - timedelta(days=days)).date().isoformat()
    end   = now.date().isoformat()
    hdr   = {'Authorization': f'Token {DEEPGRAM_KEY}'}

    async with httpx.AsyncClient(timeout=20) as client:
        u = await client.get(
            f'{DEEPGRAM_BASE}/projects/{pid}/usage/breakdown',
            params={'start': start, 'end': end}, headers=hdr,
        )
        b = await client.get(
            f'{DEEPGRAM_BASE}/projects/{pid}/balances', headers=hdr,
        )
    return {
        'project_id': pid,
        'period': {'start': start, 'end': end},
        'usage':   u.json() if u.status_code == 200 else {'error': u.text[:300]},
        'balance': b.json() if b.status_code == 200 else {'error': b.text[:300]},
    }


# ── 使用履歴（運用デバッグ・Claude Code閲覧用）──────
@app.get('/api/usage-history')
async def get_usage_history(days: int = 7, service: str = None,
                            limit: int = 200, _: str = Depends(require_admin)):
    """
    API使用履歴を時系列で返す（プロジェクト名と紐付け）
    days: 過去N日（デフォルト7）
    service: 'whisper' / 'deepgram' / 'gemini-flash' / 'gemini-pro' でフィルタ
    limit: 最大件数
    """
    from datetime import timedelta
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    sql = '''
        SELECT u.id, u.service, u.units, u.unit_type, u.endpoint,
               u.cost_usd, u.created_at,
               u.project_id, p.name AS project_name
        FROM api_usage u
        LEFT JOIN projects p ON p.id = u.project_id
        WHERE u.created_at >= ?
    '''
    args = [since]
    if service:
        sql += ' AND u.service = ?'
        args.append(service)
    sql += ' ORDER BY u.created_at DESC LIMIT ?'
    args.append(limit)

    conn = get_conn()
    rows = conn.execute(sql, args).fetchall()

    # 集計
    summary_rows = conn.execute(
        '''SELECT service, unit_type, COUNT(*) AS calls,
                  SUM(units) AS total, SUM(cost_usd) AS cost
           FROM api_usage WHERE created_at >= ?
           GROUP BY service, unit_type ORDER BY service''',
        [since]
    ).fetchall()
    conn.close()

    return {
        'since': since,
        'days': days,
        'summary': [dict(r) for r in summary_rows],
        'records': [dict(r) for r in rows],
    }


# ── プロジェクト操作ログ（プロジェクト一覧 + 解析履歴）─
@app.get('/api/project-activity')
async def get_project_activity(_: str = Depends(require_admin)):
    """全プロジェクトの活動サマリーを返す（運用閲覧用）"""
    conn = get_conn()
    rows = conn.execute('''
        SELECT p.id, p.name, p.created_at, p.status,
               (SELECT COUNT(*) FROM recordings WHERE project_id=p.id) AS recordings,
               (SELECT COUNT(*) FROM ai_transcripts WHERE project_id=p.id) AS analyses,
               (SELECT COUNT(*) FROM merged_transcripts WHERE project_id=p.id) AS merges,
               (SELECT COUNT(*) FROM qa_history WHERE project_id=p.id) AS qa_count,
               (SELECT SUM(cost_usd) FROM api_usage WHERE project_id=p.id) AS cost_usd
        FROM projects p
        ORDER BY p.created_at DESC
    ''').fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── APIキー更新（.envを書き換え＋プロセス内反映）─────
@app.patch('/api/settings/keys')
async def update_api_keys(request: Request, _: str = Depends(require_admin)):
    """
    APIキーを更新。Body: { "whisper": "sk-...", "deepgram": "...", "gemini": "..." }
    指定されたキーのみ更新（空文字や省略は無視）。
    .envファイルを書き換え、プロセス内の変数も即時更新。
    """
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(400, 'body must be a JSON object')

    # キー名 → .envの変数名
    KEY_MAP = {
        'whisper':   'OPENAI_API_KEY',
        'deepgram':  'DEEPGRAM_API_KEY',
        'gemini':    'GEMINI_API_KEY',
        'anthropic': 'ANTHROPIC_API_KEY',
    }
    updates = {}
    for k, env_name in KEY_MAP.items():
        v = (body.get(k) or '').strip()
        if v and not v.startswith('*'):  # マスク表示そのまま送られたら無視
            updates[env_name] = v

    if not updates:
        raise HTTPException(400, 'no valid keys to update')

    # .envを読んで更新（既存行を置換 or 末尾追加）
    env_path = '/opt/jizo-api/.env'
    try:
        with open(env_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except FileNotFoundError:
        lines = []

    updated_keys = set()
    new_lines = []
    for line in lines:
        stripped = line.strip()
        # コメント・空行はそのまま
        if not stripped or stripped.startswith('#'):
            new_lines.append(line)
            continue
        if '=' in stripped:
            key_name = stripped.split('=', 1)[0].strip()
            if key_name in updates:
                new_lines.append(f'{key_name}={updates[key_name]}\n')
                updated_keys.add(key_name)
                continue
        new_lines.append(line)

    # 未追記のものは末尾に追加
    for key_name, value in updates.items():
        if key_name not in updated_keys:
            if new_lines and not new_lines[-1].endswith('\n'):
                new_lines.append('\n')
            new_lines.append(f'{key_name}={value}\n')

    with open(env_path, 'w', encoding='utf-8') as f:
        f.writelines(new_lines)

    # プロセス内変数を更新（再起動なしで反映）
    global OPENAI_KEY, DEEPGRAM_KEY, GEMINI_KEY, ANTHROPIC_KEY, _dg_project_id
    if 'OPENAI_API_KEY' in updates:
        OPENAI_KEY = updates['OPENAI_API_KEY']
        os.environ['OPENAI_API_KEY'] = OPENAI_KEY
    if 'DEEPGRAM_API_KEY' in updates:
        DEEPGRAM_KEY = updates['DEEPGRAM_API_KEY']
        os.environ['DEEPGRAM_API_KEY'] = DEEPGRAM_KEY
        _dg_project_id = None  # キー変更時はproject_idキャッシュをリセット
    if 'GEMINI_API_KEY' in updates:
        GEMINI_KEY = updates['GEMINI_API_KEY']
        os.environ['GEMINI_API_KEY'] = GEMINI_KEY
    if 'ANTHROPIC_API_KEY' in updates:
        ANTHROPIC_KEY = updates['ANTHROPIC_API_KEY']
        os.environ['ANTHROPIC_API_KEY'] = ANTHROPIC_KEY

    logger.info(f'API keys updated: {list(updates.keys())}')
    return {'ok': True, 'updated': list(updates.keys())}


@app.get('/api/settings')
async def get_settings(_: str = Depends(require_admin)):
    """APIキー（マスク済み）・上限・現在の使用量を返す"""
    services = ['whisper', 'deepgram', 'gemini-flash', 'gemini-pro', 'claude-haiku']
    keys = {
        'whisper':      {'configured': bool(OPENAI_KEY),    'masked': _mask_key(OPENAI_KEY)},
        'deepgram':     {'configured': bool(DEEPGRAM_KEY),  'masked': _mask_key(DEEPGRAM_KEY)},
        'gemini-flash': {'configured': bool(GEMINI_KEY),    'masked': _mask_key(GEMINI_KEY)},
        'gemini-pro':   {'configured': bool(GEMINI_KEY),    'masked': _mask_key(GEMINI_KEY)},
        'claude-haiku': {'configured': bool(ANTHROPIC_KEY), 'masked': _mask_key(ANTHROPIC_KEY)},
    }
    usage = {}
    for svc in services:
        daily   = _get_usage(svc, 'day')
        monthly = _get_usage(svc, 'month')
        limits  = USAGE_LIMITS.get(svc, {})
        usage[svc] = {
            'limits':  limits,
            'daily':   daily,
            'monthly': monthly,
        }
    return {'keys': keys, 'usage': usage}


# ── 解析エンジン（Whisper / Deepgram）─────────────
DEFAULT_SETTINGS = {
    1: {  # Run1 デフォルト：Whisper・日本語固定・プロンプトなし
        'engine': 'whisper',
        'language': 'ja',
        'prompt': '',
        'model': 'whisper-1',
        'diarize': False,
    },
    2: {  # Run2 デフォルト：Deepgram・日本語・話者分離on
        'engine': 'deepgram',
        'language': 'ja',
        'prompt': '',
        'model': 'nova-3',
        'diarize': True,
    },
}


async def _run_whisper(audio_path: str, settings: dict, pnames: list, keyterms: list,
                       project_id: str = None) -> tuple:
    """Whisper API呼び出し。(utterances, full_text) を返す"""
    oai = AsyncOpenAI(api_key=OPENAI_KEY)
    with open(audio_path, 'rb') as fp:
        kwargs = {
            'model': settings.get('model') or 'whisper-1',
            'file': fp,
            'response_format': 'verbose_json',
            'timestamp_granularities': ['segment'],
        }
        lang = settings.get('language') or ''
        if lang and lang != 'auto':
            kwargs['language'] = lang
        prompt = (settings.get('prompt') or '').strip()
        if prompt:
            kwargs['prompt'] = prompt
        transcript = await oai.audio.transcriptions.create(**kwargs)
    segs = getattr(transcript, 'segments', None) or []
    utterances = [
        {'speaker': 'A',
         'start': int(s.start * 1000),
         'end':   int(s.end   * 1000),
         'text':  s.text.strip()}
        for s in segs
    ]
    # 使用量記録（duration秒→分）
    duration_sec = float(getattr(transcript, 'duration', 0) or 0)
    minutes = duration_sec / 60.0
    cost = minutes * USAGE_LIMITS['whisper']['rate_usd_per_min']
    _record_usage('whisper', minutes, 'minutes', project_id, '/analyze', cost)
    return utterances, (transcript.text or ''), duration_sec


async def _run_deepgram(audio_path: str, settings: dict, pnames: list, keyterms: list,
                        project_id: str = None) -> tuple:
    """Deepgram API呼び出し。(utterances, full_text) を返す"""
    if not DEEPGRAM_KEY:
        raise HTTPException(500, 'DEEPGRAM_API_KEY not configured')

    params = {
        'model': settings.get('model') or 'nova-3',
        'language': settings.get('language') or 'ja',
        'smart_format': 'true',
        'punctuate': 'true',
        'paragraphs': 'true',
        'utterances': 'true',
    }
    if settings.get('diarize'):
        params['diarize'] = 'true'
    # 固有名詞・参加者名を補強（Nova-3は keyterm、それ以前は keywords）
    kws = list(set([*pnames, *(keyterms or [])]))[:20]
    if kws:
        model_name = (settings.get('model') or 'nova-3').lower()
        if model_name.startswith('nova-3'):
            params['keyterm'] = kws
        else:
            params['keywords'] = kws
    prompt = (settings.get('prompt') or '').strip()
    if prompt:
        # Deepgramは直接的なpromptはないが、keytermsで代用済み。promptはログだけ
        logger.info(f'Deepgram prompt (info only): {prompt[:80]}')

    with open(audio_path, 'rb') as fp:
        audio_bytes = fp.read()

    async with httpx.AsyncClient(timeout=300) as client:
        r = await client.post(
            f'{DEEPGRAM_BASE}/listen',
            params=params,
            headers={
                'Authorization': f'Token {DEEPGRAM_KEY}',
                'content-type': 'audio/webm',
            },
            content=audio_bytes,
        )
        if r.status_code != 200:
            raise HTTPException(502, f'Deepgram failed: {r.text[:500]}')
        data = r.json()

    # utterances（話者分離あり）優先、なければwordsから合成
    utts_raw = (data.get('results', {}).get('utterances') or [])
    if utts_raw:
        utterances = [
            {
                'speaker': chr(65 + (u.get('speaker') or 0)),  # 0→'A', 1→'B'…
                'start': int((u.get('start') or 0) * 1000),
                'end':   int((u.get('end')   or 0) * 1000),
                'text':  (u.get('transcript') or '').strip(),
            }
            for u in utts_raw
        ]
    else:
        # フォールバック：channels[0].alternatives[0].paragraphs
        utterances = []
        channels = data.get('results', {}).get('channels') or []
        if channels and channels[0].get('alternatives'):
            alt = channels[0]['alternatives'][0]
            for para in (alt.get('paragraphs', {}).get('paragraphs') or []):
                for s in para.get('sentences', []):
                    utterances.append({
                        'speaker': chr(65 + (para.get('speaker') or 0)),
                        'start': int((s.get('start') or 0) * 1000),
                        'end':   int((s.get('end')   or 0) * 1000),
                        'text':  (s.get('text') or '').strip(),
                    })

    full_text = ''
    chans = data.get('results', {}).get('channels') or []
    if chans and chans[0].get('alternatives'):
        full_text = chans[0]['alternatives'][0].get('transcript', '')

    # 使用量記録
    duration_sec = float(data.get('metadata', {}).get('duration') or 0)
    minutes = duration_sec / 60.0
    cost = minutes * USAGE_LIMITS['deepgram']['rate_usd_per_min']
    _record_usage('deepgram', minutes, 'minutes', project_id, '/analyze', cost)
    return utterances, full_text, duration_sec


@app.post('/api/projects/{project_id}/analyze')
async def analyze_project(project_id: str, run: int = 1, force: bool = False,
                          request: Request = None):
    """
    Body (optional JSON): { engine, language, prompt, model, diarize }
    指定なければ DEFAULT_SETTINGS[run] を使用
    """
    if run not in (1, 2):
        raise HTTPException(400, 'run must be 1 or 2')

    # 設定マージ
    settings = dict(DEFAULT_SETTINGS[run])
    try:
        body = await request.json() if request else {}
        if isinstance(body, dict):
            for k in ('engine', 'language', 'prompt', 'model', 'diarize'):
                if k in body and body[k] is not None:
                    settings[k] = body[k]
    except Exception:
        pass

    engine = settings.get('engine') or 'whisper'
    if engine not in ('whisper', 'deepgram'):
        raise HTTPException(400, f'engine must be whisper or deepgram')

    conn = get_conn()
    existing = conn.execute(
        "SELECT COUNT(*) FROM ai_transcripts WHERE project_id=? AND run_number=? AND status NOT IN ('error')",
        (project_id, run)
    ).fetchone()[0]
    if existing > 0:
        if force:
            conn.execute('DELETE FROM ai_transcripts WHERE project_id=? AND run_number=?',
                         (project_id, run))
            conn.commit()
        else:
            conn.close()
            raise HTTPException(409, f'Run {run} already exists for this project')

    recs_all = conn.execute(
        'SELECT * FROM recordings WHERE project_id=? ORDER BY seq', (project_id,)
    ).fetchall()
    if not recs_all:
        conn.close()
        raise HTTPException(404, 'No recordings for this project')

    # 0バイト or 欠損ファイルを除外（Whisper/Deepgramが落ちるのを防ぐ）
    recs = []
    for r in recs_all:
        p = r['audio_path']
        if not os.path.exists(p) or os.path.getsize(p) == 0:
            logger.warning(f'analyze: skip empty/missing audio rec_id={r["id"]} path={p}')
            continue
        recs.append(r)
    if not recs:
        conn.close()
        raise HTTPException(400, 'All recordings are empty or missing')

    keyterms = _build_keyterms(project_id, conn)
    pnames = [r['name'] for r in conn.execute(
        'SELECT name FROM participants WHERE project_id=? AND name != ""', (project_id,)
    ).fetchall()]

    # 上限チェック（音声合計の推定。録音ファイル数で粗推定し、実際の使用量は呼び出し後に記録）
    try:
        _check_limit(engine, expected_units=0, expected_unit_type='minutes')
    except HTTPException:
        conn.close()
        raise

    total_segments = 0
    for rec in recs:
        logger.info(f'analyze run={run} engine={engine} rec={rec["id"]} settings={settings}')
        try:
            if engine == 'whisper':
                utterances, full_text, duration_sec = await _run_whisper(rec['audio_path'], settings, pnames, keyterms, project_id)
            else:
                utterances, full_text, duration_sec = await _run_deepgram(rec['audio_path'], settings, pnames, keyterms, project_id)
        except HTTPException:
            conn.close()
            raise
        except Exception as e:
            conn.close()
            logger.error(f'{engine} error: {e}')
            raise HTTPException(502, f'{engine} failed: {e}')

        logger.info(f'{engine} done segments={len(utterances)} duration={duration_sec:.2f}s')
        total_segments += len(utterances)

        # recordings.duration を秒文字列で更新（管理画面のチャンクオフセット計算に使用）
        if duration_sec > 0 and not (rec['duration'] or '').strip():
            conn.execute(
                'UPDATE recordings SET duration=? WHERE id=?',
                (f'{duration_sec:.3f}', rec['id'])
            )

        ext_id = f'{engine}:{rec["id"]}:{run}'
        conn.execute(
            '''INSERT INTO ai_transcripts
               (project_id,recording_id,run_number,aai_id,status,full_text,
                utterances_json,speaker_map_json,engine,settings_json)
               VALUES(?,?,?,?,?,?,?,?,?,?)''',
            (project_id, rec['id'], run, ext_id, 'completed',
             full_text,
             json.dumps(utterances, ensure_ascii=False),
             json.dumps({}, ensure_ascii=False),
             engine,
             json.dumps(settings, ensure_ascii=False))
        )

    status_map = {1: 'completed_run1', 2: 'completed_run2'}
    conn.execute('UPDATE projects SET status=? WHERE id=?', (status_map[run], project_id))
    conn.commit()
    conn.close()
    return {'ok': True, 'run': run, 'recordings': len(recs),
            'segments': total_segments, 'engine': engine, 'settings': settings}


# ── 解析状態ポーリング + 突合 ──────────────────────
@app.get('/api/projects/{project_id}/poll')
async def poll_project(project_id: str):
    conn = get_conn()
    txs = conn.execute(
        "SELECT * FROM ai_transcripts WHERE project_id=? AND status NOT IN ('completed','error')",
        (project_id,)
    ).fetchall()

    async with httpx.AsyncClient(timeout=15) as client:
        for tx in txs:
            # Whisper は同期完了済みのためポーリング不要
            if str(tx['aai_id']).startswith('whisper:'):
                continue
            r = await client.get(f'{AAI_BASE}/v2/transcript/{tx["aai_id"]}', headers=AAI_HDR)
            d = r.json()
            st = d.get('status')
            if st == 'completed':
                utterances  = d.get('utterances') or []
                speaker_map = _build_speaker_map(project_id, utterances, conn)
                conn.execute(
                    '''UPDATE ai_transcripts
                       SET status=?, full_text=?, utterances_json=?, speaker_map_json=?
                       WHERE id=?''',
                    ('completed', d.get('text', ''),
                     json.dumps(utterances, ensure_ascii=False),
                     json.dumps(speaker_map, ensure_ascii=False),
                     tx['id'])
                )
            elif st == 'error':
                conn.execute(
                    "UPDATE ai_transcripts SET status='error', error=? WHERE id=?",
                    (d.get('error', 'unknown'), tx['id'])
                )

    # 全件の状態を集計してprojectステータスを更新
    all_txs = conn.execute(
        'SELECT run_number, status FROM ai_transcripts WHERE project_id=?', (project_id,)
    ).fetchall()

    pending = sum(1 for t in all_txs if t['status'] not in ('completed', 'error'))
    if pending == 0 and all_txs:
        run_numbers = {t['run_number'] for t in all_txs}
        if 2 in run_numbers:
            conn.execute("UPDATE projects SET status='completed_run2' WHERE id=?", (project_id,))
        else:
            conn.execute("UPDATE projects SET status='completed_run1' WHERE id=?", (project_id,))

    conn.commit()
    proj = conn.execute('SELECT * FROM projects WHERE id=?', (project_id,)).fetchone()
    conn.close()
    return {'status': proj['status'], 'pending': pending}


def _build_speaker_map(project_id: str, utterances: list, conn) -> dict:
    """タイムスタンプ突合でAssemblyAI話者クラスタ→参加者名マッピングを生成"""
    participants = {
        row['idx']: row['name']
        for row in conn.execute(
            'SELECT idx,name FROM participants WHERE project_id=?', (project_id,)
        ).fetchall()
    }
    ref_segs = conn.execute(
        'SELECT speaker_idx,ts FROM ref_segments WHERE project_id=? AND speaker_idx IS NOT NULL',
        (project_id,)
    ).fetchall()

    def ts_to_ms(ts: str) -> int:
        parts = ts.split(':')
        try:
            return (int(parts[0]) * 60 + int(parts[1])) * 1000
        except Exception:
            return 0

    ref_list = [(ts_to_ms(r['ts']), r['speaker_idx']) for r in ref_segs]

    votes: dict = {}
    for utt in utterances:
        spk   = utt.get('speaker', 'A')
        start = utt.get('start', 0)
        end   = utt.get('end', 0)
        mid   = (start + end) / 2
        votes.setdefault(spk, {})
        for (ref_ms, ref_idx) in ref_list:
            if abs(ref_ms - mid) < 10000:
                votes[spk][ref_idx] = votes[spk].get(ref_idx, 0) + 1

    result = {}
    for spk, vote_map in votes.items():
        if vote_map:
            best_idx = max(vote_map, key=lambda k: vote_map[k])
            result[spk] = participants.get(best_idx, f'話者{best_idx + 1}')
        else:
            result[spk] = f'Speaker {spk}'
    return result


# ── LLM 使用可能モデル ──────────────────────────────
# ALLOWED_MODELS は merge_project 専用（Gemini限定・JSONLプロンプトがGemini向けに最適化）
ALLOWED_MODELS = {
    'gemini-2.5-flash':  'Gemini 2.5 Flash（推奨）',
    'gemini-2.5-pro':    'Gemini 2.5 Pro（高品質）',
}
# LLM_MODELS は summarize / qa 用（Gemini + Claude）
LLM_MODELS = {
    **ALLOWED_MODELS,
    'claude-haiku-4-5-20251001': 'Claude Haiku 4.5（要約・Q&A向け）',
}


# ── LLM統合：設定範囲と取得 ──────────────────────────
MERGE_SETTINGS_RANGES = {
    'chunk_size':           (int,   10, 50),
    'cluster_window_ms':    (int,   500, 5000),
    'orphan_sim_threshold': (float, 0.0, 1.0),
    'parallel':             (int,   1, 5),
    'max_tokens_per_chunk': (int,   4096, 16384),
    'retry_per_chunk':      (int,   0, 3),
    'default_w_ws':         (int,   0, 10),
    'default_w_run1':       (int,   0, 10),
    'default_w_run2':       (int,   0, 10),
}
MERGE_SETTINGS_DEFAULTS = {
    'chunk_size':           25,
    'cluster_window_ms':    2000,
    'orphan_sim_threshold': 0.4,
    'backbone_algo':        'fixed',
    'backbone_fixed':       'run2',
    'default_model':        'gemini-2.5-flash',
    'parallel':             3,
    'max_tokens_per_chunk': 8192,
    'retry_per_chunk':      1,
    'speaker_priority':     'deepgram',
    'default_w_ws':         0,
    'default_w_run1':       7,
    'default_w_run2':       3,
}


def _get_merge_settings() -> dict:
    """DBから設定読出。未存在キーはデフォルト。"""
    try:
        conn = get_conn()
        rows = conn.execute('SELECT key,value FROM merge_settings').fetchall()
        conn.close()
        raw = {r['key']: r['value'] for r in rows}
    except Exception:
        raw = {}
    out = dict(MERGE_SETTINGS_DEFAULTS)
    for k, v in raw.items():
        if k not in MERGE_SETTINGS_DEFAULTS:
            continue
        if k in MERGE_SETTINGS_RANGES:
            t, lo, hi = MERGE_SETTINGS_RANGES[k]
            try:
                val = t(v)
                out[k] = max(lo, min(hi, val))
            except Exception:
                pass
        else:
            out[k] = v
    return out


def _clamp_merge_settings(d: dict) -> dict:
    """入力値を範囲クランプ。enum・文字列は許可リスト判定。"""
    out = {}
    for k, default in MERGE_SETTINGS_DEFAULTS.items():
        if k not in d:
            out[k] = default
            continue
        v = d[k]
        if k in MERGE_SETTINGS_RANGES:
            t, lo, hi = MERGE_SETTINGS_RANGES[k]
            try:
                out[k] = max(lo, min(hi, t(v)))
            except Exception:
                out[k] = default
        elif k == 'backbone_algo':
            out[k] = v if v in ('rows_x_log_chars', 'rows_only', 'chars_only', 'fixed') else default
        elif k == 'backbone_fixed':
            out[k] = v if v in ('', 'ws', 'run1', 'run2') else ''
        elif k == 'speaker_priority':
            out[k] = v if v in ('deepgram', 'ws', 'hybrid') else default
        elif k == 'default_model':
            out[k] = v if v in ALLOWED_MODELS else default
        else:
            out[k] = v
    return out


def _normalize_for_compare(text: str) -> str:
    """類似度比較・幻覚検出用の正規化（句読点・スペース除去・要確認タグ除去）"""
    if not text:
        return ''
    t = re.sub(r'([^\x00-\x7F])\s+([^\x00-\x7F])', r'\1\2', text)
    t = re.sub(r'([^\x00-\x7F])\s+([^\x00-\x7F])', r'\1\2', t)
    t = re.sub(r'【要確認[^】]*】', '', t)
    return re.sub(r'[。、！？!?,\.・\s]', '', t)


def _text_sim(a: str, b: str) -> float:
    """0..1。短文の包含優先 + Dice 係数（bigram）"""
    na, nb = _normalize_for_compare(a), _normalize_for_compare(b)
    if not na or not nb:
        return 0.0
    short, long_ = (na, nb) if len(na) <= len(nb) else (nb, na)
    if len(short) >= 2 and short in long_:
        return len(short) / len(long_)
    if len(na) < 2 or len(nb) < 2:
        return 1.0 if na == nb else 0.0
    def grams(s):
        return {s[i:i+2] for i in range(len(s) - 1)}
    A, B = grams(na), grams(nb)
    if len(A) + len(B) == 0:
        return 0.0
    return 2 * len(A & B) / (len(A) + len(B))


def _select_backbone(ws_items, run1_items, run2_items, algo: str, fixed: str):
    """主軸選定。returns (key, items) or raises HTTPException."""
    import math
    sources = {
        'ws':   ws_items,
        'run1': run1_items,
        'run2': run2_items,
    }
    if algo == 'fixed' and fixed in sources and sources[fixed]:
        return fixed, sources[fixed]
    def score(items):
        n = len(items)
        if n == 0:
            return -1.0
        total_chars = sum(len(x.get('text', '')) for x in items)
        if algo == 'rows_only':
            return float(n)
        if algo == 'chars_only':
            return float(total_chars)
        return n * math.log(total_chars + 1)
    candidates = [(k, score(v)) for k, v in sources.items() if v]
    if not candidates:
        raise HTTPException(400, 'マージ可能なソースがありません')
    candidates.sort(key=lambda x: -x[1])
    best_key = candidates[0][0]
    return best_key, sources[best_key]


def _build_rows_with_orphans(backbone_key: str, backbone_items: list,
                              ws_items: list, run1_items: list, run2_items: list,
                              window_ms: int, sim_threshold: float):
    """主軸の各行 + 主軸窓外のオーファンを時系列で構築。"""
    others = {
        'ws':   ws_items   if backbone_key != 'ws'   else [],
        'run1': run1_items if backbone_key != 'run1' else [],
        'run2': run2_items if backbone_key != 'run2' else [],
    }

    def nearest(items, ms):
        if not items:
            return None
        best = min(items, key=lambda x: abs(x['ms'] - ms))
        return best if abs(best['ms'] - ms) <= window_ms else None

    rows = []
    for item in backbone_items:
        ms = item['ms']
        cand_ws = item if backbone_key == 'ws' else nearest(ws_items, ms)
        cand_r1 = item if backbone_key == 'run1' else nearest(run1_items, ms)
        cand_r2 = item if backbone_key == 'run2' else nearest(run2_items, ms)
        rows.append({
            'ms': ms,
            'ts': item.get('ts', ''),
            'backbone_text': (item.get('text') or '').strip(),
            'backbone_src': backbone_key,
            'cand_ws':   (cand_ws['text']  if cand_ws else ''),
            'cand_run1': (cand_r1['text']  if cand_r1 else ''),
            'cand_run2': (cand_r2['text']  if cand_r2 else ''),
            'is_orphan': False,
        })

    # オーファン検出
    orphan_items = []
    for src_key, items in others.items():
        for it in items:
            ms = it['ms']
            nearest_backbone = min(rows, key=lambda r: abs(r['ms'] - ms)) if rows else None
            if nearest_backbone and abs(nearest_backbone['ms'] - ms) <= window_ms:
                continue
            # 広域でテキスト類似な主軸行があれば skip
            sim_high = False
            for r in rows:
                if abs(r['ms'] - ms) <= window_ms * 3:
                    if _text_sim(it.get('text', ''), r['backbone_text']) >= sim_threshold:
                        sim_high = True
                        break
            if sim_high:
                continue
            orphan_items.append((ms, src_key, it))

    # 同一ソース連続オーファン連結（500ms以内）
    orphan_items.sort(key=lambda x: x[0])
    merged_orphans = []
    last = None
    for ms, src, it in orphan_items:
        text = (it.get('text') or '').strip()
        if last and last['src'] == src and ms - last['ms_last'] <= 500:
            last['text'] = (last['text'] + ' ' + text).strip()
            last['ms_last'] = ms
        else:
            last = {'ms': ms, 'ms_last': ms, 'src': src, 'text': text}
            merged_orphans.append(last)

    def _ms_to_ts_local(ms):
        sec = max(0, int(ms / 1000))
        return f'{sec // 60:02d}:{sec % 60:02d}'

    for o in merged_orphans:
        rows.append({
            'ms': o['ms'],
            'ts': _ms_to_ts_local(o['ms']),
            'backbone_text': o['text'],
            'backbone_src': o['src'],
            'cand_ws':   o['text'] if o['src'] == 'ws'   else '',
            'cand_run1': o['text'] if o['src'] == 'run1' else '',
            'cand_run2': o['text'] if o['src'] == 'run2' else '',
            'is_orphan': True,
        })

    rows.sort(key=lambda r: r['ms'])
    for i, r in enumerate(rows):
        r['idx'] = i + 1
    return rows


def _dg_speaker_at(ms, run2_items, name_map, window_ms=3000):
    """Run2 (Deepgram) の最近傍 utterance から話者を解決。
    speaker は 'A'/'B'/... または 0/1/2 を想定。
    name_map (参加者 idx→名前) があれば優先マッピング、なければ 'Speaker X' 返却。"""
    if not run2_items:
        return None
    nearest = min(run2_items, key=lambda x: abs(x['ms'] - ms))
    if abs(nearest['ms'] - ms) > window_ms:
        return None
    spk = nearest.get('speaker')
    if spk is None or spk == '':
        return None
    try:
        if isinstance(spk, str) and len(spk) == 1 and spk.upper().isalpha():
            idx = ord(spk.upper()) - ord('A')
        else:
            idx = int(spk)
        if idx in name_map:
            return name_map[idx]
        return f'Speaker {spk}'
    except Exception:
        return f'Speaker {spk}'


def _resolve_speaker_for_row(ms, sp_events_sorted, ws_items, name_map,
                              run2_items=None, priority='ws'):
    """priority: 'deepgram' / 'ws' / 'hybrid'。run2_items は Phase 2 で集めた Run2 utterance 群。"""
    def _ws_resolve():
        near = [e for e in sp_events_sorted if abs(e['ms'] - ms) <= 2000]
        if near:
            best = min(near, key=lambda e: abs(e['ms'] - ms))
            return (best['speaker_name'] or '未設定', False)
        prior = [e for e in sp_events_sorted if e['ms'] <= ms]
        if prior:
            return (prior[-1]['speaker_name'] or '未設定', False)
        if ws_items:
            nearest_ws = min(ws_items, key=lambda x: abs(x['ms'] - ms))
            if abs(nearest_ws['ms'] - ms) <= 4000 and nearest_ws['speaker_idx'] is not None:
                return (name_map.get(nearest_ws['speaker_idx'], '未設定'), False)
        return ('未設定', True)

    if priority == 'deepgram':
        dg = _dg_speaker_at(ms, run2_items, name_map)
        if dg:
            return (dg, False)
        return _ws_resolve()
    if priority == 'hybrid':
        near = [e for e in sp_events_sorted if abs(e['ms'] - ms) <= 2000]
        if near:
            best = min(near, key=lambda e: abs(e['ms'] - ms))
            return (best['speaker_name'] or '未設定', False)
        dg = _dg_speaker_at(ms, run2_items, name_map)
        if dg:
            return (dg, False)
        return _ws_resolve()
    return _ws_resolve()


def _chunk_rows(rows: list, size: int, ctx: int = 2):
    chunks = []
    n = len(rows)
    for s in range(0, n, size):
        e = min(s + size, n)
        chunks.append({
            'idx_start': s,
            'idx_end':   e,
            'rows':      rows[s:e],
            'ctx_prev':  rows[max(0, s - ctx):s],
            'ctx_next':  rows[e:min(n, e + ctx)],
        })
    return chunks


def _detect_hallucination(text: str, candidates: list, ratio_threshold: float = 0.5) -> bool:
    """text の4文字窓のうち、候補連結文字列に存在しない割合が ratio_threshold を超えたら
    「過剰補填の可能性あり」とみなす。漢字変換違い・小さな自然化（候補外が散発）は許容し、
    丸ごとの捏造（候補外が過半数）のみ捕捉する。"""
    norm = _normalize_for_compare(text)
    joined = ''.join(_normalize_for_compare(c or '') for c in candidates)
    if not norm or not joined or len(norm) < 4:
        return False
    total = len(norm) - 3
    novel = sum(1 for i in range(total) if norm[i:i+4] not in joined)
    return (novel / total) > ratio_threshold


def _parse_jsonl(raw: str, expected_n: int):
    """JSONL or JSON array をパース。idx が 1..expected_n を各1個ずつ網羅していれば
    idx 昇順のリストを返す。欠落/重複/余剰があれば None（=主軸フォールバック）。
    各オブジェクトの "text" は文字列 or 配列（文分割）を許容する。"""
    raw = raw.strip()
    fence_match = re.search(r'```(?:json)?\s*(.*?)(?:```|\Z)', raw, re.DOTALL)
    if fence_match:
        raw = fence_match.group(1).strip()

    items = []
    if raw.startswith('['):
        try:
            data = json.loads(raw)
            if isinstance(data, list):
                items = [d for d in data if isinstance(d, dict)]
        except Exception:
            items = []
    if not items:
        for line in raw.split('\n'):
            line = line.strip().rstrip(',')
            if not line or line in ('[', ']'):
                continue
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    items.append(obj)
            except Exception:
                continue

    # idx 網羅検証：1..expected_n を各1個ずつ
    by_idx = {}
    for obj in items:
        idx = obj.get('idx')
        if not isinstance(idx, int) or idx < 1 or idx > expected_n:
            continue
        if idx in by_idx:  # 重複 idx は不正
            return None
        by_idx[idx] = obj
    if len(by_idx) != expected_n:
        return None
    return [by_idx[i] for i in range(1, expected_n + 1)]


async def _call_gemini_chunk(system_prompt: str, user_prompt: str,
                              model: str, max_tokens: int, retry: int = 1):
    """単一チャンクのLLM呼出。失敗時は None。"""
    last_err = None
    for attempt in range(retry + 1):
        try:
            payload = {
                'model': model,
                'messages': [
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user',   'content': user_prompt},
                ],
                'max_tokens': max_tokens,
            }
            async with httpx.AsyncClient(timeout=90) as client:
                r = await client.post(
                    f'{GEMINI_BASE}/chat/completions',
                    headers={
                        'Authorization': f'Bearer {GEMINI_KEY}',
                        'content-type': 'application/json',
                    },
                    json=payload,
                )
                if r.status_code != 200:
                    last_err = f'status={r.status_code} body={r.text[:200]}'
                    continue
                return r.json()['choices'][0]['message']['content'].strip()
        except Exception as e:
            last_err = str(e)
    logger.warning(f'_call_gemini_chunk failed after {retry+1} attempts: {last_err}')
    return None


def _build_chunk_prompt(chunk: dict, plist: str, w_ws: int, w_run1: int, w_run2: int,
                        run_engines: dict):
    system_prompt = (
        'あなたは日本語会議録音の文字起こしを整流する編集者です。\n'
        '\n'
        'あなたの仕事は、主軸ソースの行ごとのテキストに対して、'
        '他ソース（補完候補）を文脈として参照しながら以下のみを行うことです：\n'
        '1. 主軸が拾い損ねた単語の補完\n'
        '2. 主軸内で明らかに順序が崩れた語句の修正\n'
        '3. 句読点とスペースの整形\n'
        '4. 隣接レコード境界の重複除去：直前/直後の行（「文脈」に表示済み）と同一発話の断片が'
        '当該行の先頭または末尾に重複している場合（漢字変換違い・区切り違い・送り仮名違いを含む同一発話）、'
        '当該行から重複部分のみをトリムして自然な境界にする。'
        '明確な重複のみトリムし、別発話か判断に迷う場合は残す。\n'
        '5. 文脈による文分割：1つの行のテキストが句読点が無くても明らかに2文以上に分かれている場合、'
        '"text" を文単位の配列にして分割してよい（例：「絶句してましたそれはそうだろう」→'
        '["絶句してました","それはそうだろう"]）。分割は再区切りのみで、文字の追加・削除・新規生成は禁止'
        '（配列を連結すると元の内容と一致すること）。1文なら文字列のまま返す。明らかに2文以上のときだけ分割し、迷えば分割しない。\n'
        '\n'
        '絶対禁止事項：\n'
        '- 行(idx)の追加・削除・統合・並べ替え（ただし1行の "text" を文単位で配列分割することは上記5で許可）\n'
        '- 候補テキストに存在しない語句の新規生成\n'
        '- 主軸テキストの意味を変える書換え\n'
        '\n'
        '## テキスト本文の優先採用ルール（重要）\n'
        '- 主軸（多くの場合 Run2 / Deepgram）はレコード分離（行の切れ目）と話者の正確性のために選定されています。\n'
        '- テキスト本文は Run1（Whisper）を基本として採用してください（Whisper はテキスト精度が高い場合が多い）。\n'
        '- ただし、ある行の Run1 候補が明らかに不自然・文意不通・日本語として破綻している場合'
        '（意味の通らないカタカナ羅列、同音異義の誤変換など）で、かつ他ソース（Run2/WS）に同一発話の'
        'より自然で意味の通る候補が存在する場合は、そのより自然な候補を本文として採用してよい。\n'
        '- 複数候補のうち一方が意味の通る日本語、他が破綻している場合は、意味の通る方を優先する。\n'
        '- Run2 のテキストは「Run1 候補がない箇所の補完」「Run1 と意味が一致する整合確認」'
        'に加え「Run1 が破綻している箇所の置換」にも使ってよい。\n'
        '- 保守的に：明らかな破綻のときのみ置換し、迷う場合は Run1 のまま'
        '（固有名詞・専門用語を誤った候補で置換しないこと）。\n'
        '- ただし行の構造（idx・順序・ts・話者）は主軸を絶対遵守し、変更してはなりません。\n'
        '\n'
        '判断に迷う場合は主軸テキストをそのまま返してください。誠実さが速度より大切です。'
    )

    def _fmt_ctx(rs, prefix):
        if not rs:
            return '  （なし）'
        return '\n'.join(
            f'  [{prefix}{i+1}] [ts={r["ts"]}] [話者={r.get("speaker","")}] 主軸: {r["backbone_text"]}'
            for i, r in enumerate(rs)
        )

    def _fmt_target(rs):
        out = []
        for i, r in enumerate(rs):
            idx = i + 1
            out.append(
                f'行{idx:03d} [ts={r["ts"]}] [話者={r.get("speaker","")}]\n'
                f'    主軸({r["backbone_src"]}): {r["backbone_text"]}\n'
                f'    WS:   {r["cand_ws"]   or "（なし）"}\n'
                f'    Run1: {r["cand_run1"] or "（なし）"}\n'
                f'    Run2: {r["cand_run2"] or "（なし）"}'
            )
        return '\n'.join(out)

    N = len(chunk['rows'])
    user_prompt = (
        f"## 参加者\n{plist}\n\n"
        f"## 補完優先度（同等候補から1つ選ぶ際の参考）\n"
        f"WebSpeech : Run1({run_engines.get(1,'ASR')}) : Run2({run_engines.get(2,'ASR')}) = "
        f"{w_ws} : {w_run1} : {w_run2}\n\n"
        f"## 文脈（参照のみ・出力に含めない）\n{_fmt_ctx(chunk['ctx_prev'], 'prev-')}\n\n"
        f"## 編集対象（必ず {N} 行を {N} 行で返す）\n{_fmt_target(chunk['rows'])}\n\n"
        f"## 文脈（参照のみ・出力に含めない）\n{_fmt_ctx(chunk['ctx_next'], 'next-')}\n\n"
        f"## 出力形式（JSONL、1行=1JSON、改行区切り、フェンス不要）\n"
        f'{{"idx":1,"text":"...","note":null}}\n'
        f'{{"idx":2,"text":["文1","文2"],"note":null}}   ← 文分割する場合は text を配列に\n'
        f'{{"idx":3,"text":"...","note":"要確認の理由"}}\n'
        f"...\n\n"
        f"※ 入力 {N} 行それぞれに対し idx 付き JSON を1つずつ、必ず {N} 個返してください"
        f"（idx は 1〜{N} の連番厳守・各idx1個）。文分割で最終的な文数が増えても、idx の個数は {N} のままです。"
    )
    return system_prompt, user_prompt


def _make_segment(row, text, note):
    return {
        'ts': row['ts'],
        'speaker': row.get('speaker', '未設定'),
        'text': text,
        'note': note,
        '_cand_ws':   row.get('cand_ws', ''),
        '_cand_run1': row.get('cand_run1', ''),
        '_cand_run2': row.get('cand_run2', ''),
    }


# ── LLM突合で最終版トランスクリプト生成（5フェーズ新設計） ──
@app.post('/api/projects/{project_id}/merge')
async def merge_project(project_id: str,
                        model: str = '',
                        w_ws: int = -1,
                        w_run1: int = -1,
                        w_run2: int = -1,
                        base: str = '',  # 後方互換用・未使用
                        _: str = Depends(require_admin)):
    settings = _get_merge_settings()
    if not model:
        model = settings['default_model']
    # 重み未指定（-1）時は設定 default_w_* を使用
    if w_ws < 0:
        w_ws = settings.get('default_w_ws', 0)
    if w_run1 < 0:
        w_run1 = settings.get('default_w_run1', 7)
    if w_run2 < 0:
        w_run2 = settings.get('default_w_run2', 3)
    if model not in ALLOWED_MODELS:
        raise HTTPException(400, f'model must be one of: {list(ALLOWED_MODELS.keys())}')

    conn = get_conn()
    proj = conn.execute('SELECT * FROM projects WHERE id=?', (project_id,)).fetchone()
    if not proj:
        conn.close()
        raise HTTPException(404, 'Project not found')
    try:
        source = (proj['source'] or 'mobile') if 'source' in proj.keys() else 'mobile'
    except Exception:
        source = 'mobile'

    participants = conn.execute(
        'SELECT idx,name FROM participants WHERE project_id=? ORDER BY idx', (project_id,)
    ).fetchall()
    ref_segs = conn.execute(
        'SELECT seq,text,speaker_idx,ts FROM ref_segments WHERE project_id=? ORDER BY seq',
        (project_id,)
    ).fetchall()
    speaker_events = conn.execute(
        'SELECT ts,ms,speaker_idx,speaker_name FROM speaker_events WHERE project_id=? ORDER BY ms',
        (project_id,)
    ).fetchall()
    name_map = {r['idx']: (r['name'] or f'話者{r["idx"]+1}') for r in participants}

    txs = conn.execute(
        "SELECT * FROM ai_transcripts WHERE project_id=? AND status='completed' ORDER BY run_number,id",
        (project_id,)
    ).fetchall()
    if source == 'zoom':
        recs_for_offset = conn.execute(
            'SELECT id,seq,duration,zoom_participant_name FROM recordings WHERE project_id=? ORDER BY seq',
            (project_id,)
        ).fetchall()
    else:
        recs_for_offset = conn.execute(
            'SELECT id,seq,duration FROM recordings WHERE project_id=? ORDER BY seq', (project_id,)
        ).fetchall()
    conn.close()

    zoom_speaker_by_rec_id = {}
    if source == 'zoom':
        for r in recs_for_offset:
            try:
                zoom_speaker_by_rec_id[r['id']] = r['zoom_participant_name'] or ''
            except (KeyError, IndexError):
                pass

    offset_by_rec_id = {}
    _cum = 0
    for r in recs_for_offset:
        offset_by_rec_id[r['id']] = _cum
        try:
            _cum += int(float(r['duration'] or 0) * 1000)
        except Exception:
            pass

    # 重み正規化（合計10）— LLMヒントとして使用
    w_ws   = max(0, int(w_ws))
    w_run1 = max(0, int(w_run1))
    w_run2 = max(0, int(w_run2))
    wsum = w_ws + w_run1 + w_run2
    if wsum == 0:
        w_ws, w_run1, w_run2 = 1, 7, 2
    elif wsum != 10:
        w_ws   = round(w_ws   * 10 / wsum)
        w_run1 = round(w_run1 * 10 / wsum)
        w_run2 = 10 - w_ws - w_run1

    plist = ', '.join(f'{r["name"]}（{r["idx"]}）' for r in participants) or '（参加者登録なし）'

    def _ms_to_ts(ms):
        sec = max(0, int(ms / 1000))
        return f'{sec // 60:02d}:{sec % 60:02d}'

    def _ts_to_ms(ts):
        try:
            mm, ss = ts.split(':')
            return (int(mm) * 60 + int(ss)) * 1000
        except Exception:
            return 0

    ws_items = [
        {'ms': _ts_to_ms(r['ts']), 'ts': r['ts'], 'text': (r['text'] or '').strip(),
         'speaker_idx': r['speaker_idx']}
        for r in ref_segs if (r['text'] or '').strip()
    ]

    run_items = {1: [], 2: []}
    run_engines = {1: 'ASR', 2: 'ASR'}
    for tx in txs:
        rn = tx['run_number']
        eng = (tx['engine'] or '').lower() if 'engine' in tx.keys() else ''
        if not eng and tx['aai_id']:
            eng = str(tx['aai_id']).split(':', 1)[0]
        run_engines[rn] = {'whisper': 'Whisper', 'deepgram': 'Deepgram'}.get(eng, eng or 'ASR')
        utts = json.loads(tx['utterances_json']) if tx['utterances_json'] else []
        base_offset = offset_by_rec_id.get(tx['recording_id'], 0)
        for u in utts:
            text = (u.get('text') or '').strip()
            if not text:
                continue
            abs_ms = int(u.get('start') or 0) + base_offset
            spk = u.get('speaker', 'A')
            if source == 'zoom':
                spk = zoom_speaker_by_rec_id.get(tx['recording_id'], spk) or spk
            run_items[rn].append({
                'ms': abs_ms,
                'ts': _ms_to_ts(abs_ms),
                'text': text,
                'speaker': spk,
            })

    # ── Phase 1: 主軸選定 ──────────────────────────
    backbone_key, backbone_items = _select_backbone(
        ws_items, run_items[1], run_items[2],
        settings['backbone_algo'], settings['backbone_fixed']
    )

    # ── Phase 2: 行構築（主軸 + オーファン）─────────
    rows = _build_rows_with_orphans(
        backbone_key, backbone_items,
        ws_items, run_items[1], run_items[2],
        window_ms=settings['cluster_window_ms'],
        sim_threshold=settings['orphan_sim_threshold'],
    )

    # ── Phase 3: 話者解決 ──────────────────────────
    sp_events_sorted = sorted(speaker_events, key=lambda e: e['ms'])
    for row in rows:
        sp_name, sp_uncertain = _resolve_speaker_for_row(
            row['ms'], sp_events_sorted, ws_items, name_map,
            run2_items=run_items[2],
            priority=settings.get('speaker_priority', 'deepgram'),
        )
        row['speaker'] = sp_name
        row['sp_uncertain'] = sp_uncertain

    logger.info(
        f'merge start project={project_id} model={model} '
        f'backbone={backbone_key} rows={len(rows)} '
        f'ws={len(ws_items)} r1={len(run_items[1])} r2={len(run_items[2])} '
        f'settings={settings}'
    )

    # ── Phase 4: チャンク分割 LLM 整流（並列）──────
    chunks = _chunk_rows(rows, settings['chunk_size'], ctx=2)
    sem = asyncio.Semaphore(settings['parallel'])

    async def _process_chunk(chunk):
        async with sem:
            sp, up = _build_chunk_prompt(chunk, plist, w_ws, w_run1, w_run2, run_engines)
            raw = await _call_gemini_chunk(
                sp, up, model,
                max_tokens=settings['max_tokens_per_chunk'],
                retry=settings['retry_per_chunk']
            )
            if raw is None:
                return None
            return _parse_jsonl(raw, len(chunk['rows']))

    llm_outputs = await asyncio.gather(*[_process_chunk(c) for c in chunks])

    # ── Phase 5: 強制整合 ──────────────────────────
    final_segs = []
    fallback_chunks = 0
    halluc_rows = 0
    for chunk, llm_out in zip(chunks, llm_outputs):
        if llm_out is None or len(llm_out) != len(chunk['rows']):
            fallback_chunks += 1
            for row in chunk['rows']:
                final_text = row['backbone_text']
                if row.get('sp_uncertain') and '【要確認:話者】' not in final_text:
                    final_text = (final_text + ' 【要確認:話者】').strip()
                final_segs.append(_make_segment(row, final_text, None))
        else:
            for in_row, out in zip(chunk['rows'], llm_out):
                raw_text = out.get('text') if isinstance(out, dict) else ''
                note = (out.get('note') if isinstance(out, dict) else None)
                candidates = [in_row.get('cand_ws',''), in_row.get('cand_run1',''),
                              in_row.get('cand_run2',''), in_row['backbone_text']]

                # text は文字列 or 配列（文分割）
                if isinstance(raw_text, list):
                    parts = [str(p).strip() for p in raw_text if str(p).strip()]
                else:
                    s = str(raw_text or '').strip()
                    parts = [s] if s else []

                if not parts:
                    # 空出力のみ主軸フォールバック（戻すべきLLMテキストが無いため）
                    final_text = in_row['backbone_text']
                    if in_row.get('sp_uncertain') and '【要確認:話者】' not in final_text:
                        final_text = (final_text + ' 【要確認:話者】').strip()
                    final_segs.append(_make_segment(in_row, final_text, note))
                    continue

                # 過剰補填の可能性は note で警告するが、LLMテキストは残す（garbled主軸に戻さない）。
                # 分割は文字を増やさないため、連結テキストで1回判定。
                if _detect_hallucination(''.join(parts), candidates):
                    halluc_rows += 1
                    if not note:
                        note = '要確認:LLM過剰補填可能性'

                # 分割 parts を各セグメント化（ts・話者・候補は親行を継承）
                for j, part in enumerate(parts):
                    seg_text = part
                    seg_note = note if j == 0 else None       # note は先頭セグメントのみ
                    # 話者不確実マーカーは最後のセグメントにのみ付与
                    if (j == len(parts) - 1 and in_row.get('sp_uncertain')
                            and '【要確認:話者】' not in seg_text):
                        seg_text = (seg_text + ' 【要確認:話者】').strip()
                    final_segs.append(_make_segment(in_row, seg_text, seg_note))

    logger.info(
        f'merge llm done chunks={len(chunks)} fallback={fallback_chunks} '
        f'hallucinations={halluc_rows} final_rows={len(final_segs)}'
    )

    # ── sources フィールド付加（Excel・管理画面用）─
    for seg in final_segs:
        sources = {}
        norm_final = _normalize_for_compare(seg['text'])
        for label, key in (('Web Speech', '_cand_ws'), ('Run1', '_cand_run1'), ('Run2', '_cand_run2')):
            v = seg.get(key, '')
            if v and _normalize_for_compare(v) != norm_final:
                sources[label] = v
        if sources:
            seg['sources'] = sources
        for k in ('_cand_ws', '_cand_run1', '_cand_run2'):
            seg.pop(k, None)

    notes = [
        {'ts': seg.get('ts'), 'speaker': seg.get('speaker'), 'note': seg.get('note')}
        for seg in final_segs if seg.get('note')
    ]

    conn = get_conn()
    conn.execute(
        'INSERT INTO merged_transcripts(project_id,model,result_json,notes_json,created_at) VALUES(?,?,?,?,?)',
        (project_id, model,
         json.dumps(final_segs, ensure_ascii=False),
         json.dumps(notes, ensure_ascii=False),
         now_iso())
    )
    conn.execute("UPDATE projects SET status='merged' WHERE id=?", (project_id,))
    conn.commit()
    conn.close()

    return {
        'ok': True,
        'segments': len(final_segs),
        'notes': len(notes),
        'backbone': backbone_key,
        'fallback_chunks': fallback_chunks,
        'hallucinations': halluc_rows,
    }


# ── マージ設定 API ─────────────────────────────────
@app.get('/api/settings/merge')
def get_merge_settings_endpoint(_: str = Depends(require_admin)):
    return _get_merge_settings()


@app.put('/api/settings/merge')
async def update_merge_settings_endpoint(request: Request, _: str = Depends(require_admin)):
    body = await request.json()
    clamped = _clamp_merge_settings(body or {})
    conn = get_conn()
    now = now_iso()
    for k, v in clamped.items():
        conn.execute(
            'INSERT INTO merge_settings(key,value,updated_at) VALUES(?,?,?) '
            'ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at',
            (k, str(v), now)
        )
    conn.commit()
    conn.close()
    return _get_merge_settings()


# ── 共通：Gemini呼び出しヘルパー ──────────────────
async def _call_claude(system_prompt: str, user_prompt: str,
                       model: str = 'claude-haiku-4-5-20251001',
                       max_tokens: int = 4096, project_id: str = None) -> str:
    """Anthropic Messages API 呼び出し。Gemini 互換のテキスト応答を返す。"""
    if not ANTHROPIC_KEY:
        raise HTTPException(503, 'Anthropic API key is not configured')
    svc = _service_key(model)
    _check_limit(svc, expected_units=1, expected_unit_type='requests')

    payload = {
        'model': model,
        'system': system_prompt,
        'messages': [{'role': 'user', 'content': user_prompt}],
        'max_tokens': max_tokens,
    }
    async with httpx.AsyncClient(timeout=120) as client:
        r = await client.post(
            f'{ANTHROPIC_BASE}/messages',
            headers={
                'x-api-key': ANTHROPIC_KEY,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json',
            },
            json=payload,
        )
        if r.status_code != 200:
            logger.error(f'Anthropic error: {r.text[:500]}')
            raise HTTPException(502, f'Anthropic API failed: {r.text}')
        data = r.json()

    # 使用量記録
    usage = data.get('usage') or {}
    pt = int(usage.get('input_tokens')  or 0)
    ct = int(usage.get('output_tokens') or 0)
    total = pt + ct
    if total > 0:
        lim = USAGE_LIMITS.get(svc, {})
        cost = (pt / 1_000_000) * lim.get('rate_usd_per_1m_in', 0) \
             + (ct / 1_000_000) * lim.get('rate_usd_per_1m_out', 0)
        _record_usage(svc, total, 'tokens', project_id, '/llm', cost)

    parts = data.get('content') or []
    return ''.join(p.get('text', '') for p in parts if p.get('type') == 'text').strip()


async def _call_llm(system_prompt: str, user_prompt: str, model: str,
                    max_tokens: int = 4096, project_id: str = None) -> str:
    """モデル名で Gemini / Claude を振り分ける汎用 LLM 呼び出し。"""
    if model not in LLM_MODELS:
        raise HTTPException(400, f'model must be one of: {list(LLM_MODELS.keys())}')
    if model.startswith('claude'):
        return await _call_claude(system_prompt, user_prompt, model, max_tokens, project_id)
    return await _call_gemini(system_prompt, user_prompt, model, max_tokens, project_id)


async def _call_gemini(system_prompt: str, user_prompt: str, model: str = 'gemini-2.5-flash',
                       max_tokens: int = 4096, project_id: str = None) -> str:
    if model not in ALLOWED_MODELS:
        raise HTTPException(400, f'model must be one of: {list(ALLOWED_MODELS.keys())}')
    svc = _service_key(model)
    _check_limit(svc, expected_units=1, expected_unit_type='requests')

    payload = {
        'model': model,
        'messages': [
            {'role': 'system', 'content': system_prompt},
            {'role': 'user',   'content': user_prompt},
        ],
        'max_tokens': max_tokens,
    }
    async with httpx.AsyncClient(timeout=120) as client:
        r = await client.post(
            f'{GEMINI_BASE}/chat/completions',
            headers={
                'Authorization': f'Bearer {GEMINI_KEY}',
                'content-type': 'application/json',
            },
            json=payload,
        )
        if r.status_code != 200:
            logger.error(f'Gemini error: {r.text[:500]}')
            raise HTTPException(502, f'Gemini API failed: {r.text}')
        data = r.json()

    # 使用量記録
    usage = data.get('usage') or {}
    pt = int(usage.get('prompt_tokens')     or 0)
    ct = int(usage.get('completion_tokens') or 0)
    total = pt + ct
    if total > 0:
        lim = USAGE_LIMITS.get(svc, {})
        cost = (pt / 1_000_000) * lim.get('rate_usd_per_1m_in', 0) \
             + (ct / 1_000_000) * lim.get('rate_usd_per_1m_out', 0)
        _record_usage(svc, total, 'tokens', project_id, '/llm', cost)

    return data['choices'][0]['message']['content'].strip()


# ── 用途別テンプレート（要約・Q&A のプロンプト文言切替）──────
SUMMARY_TEMPLATES = {
    'meeting': {
        'label': '会議',
        'tasks_label': '残タスク',
        'summary_role': 'あなたは日本語会議分析の専門家です。',
        'summary_focus': '決定事項・議論の流れ・結論',
        'tasks_focus': '残タスク／アクションアイテムを1つずつ箇条書きで。担当者がいれば「[名前] 〇〇する」形式',
        'qa_role': 'あなたは会議の議事録から質問に答える専門家です。',
        'source_label': '会議',
    },
    'interview': {
        'label': 'インタビュー',
        'tasks_label': 'フォローアップ事項',
        'summary_role': 'あなたは日本語インタビュー分析の専門家です。',
        'summary_focus': 'インタビュイーの発言要旨・主要トピック・重要な見解',
        'tasks_focus': 'フォローアップ事項・追加で確認したい点・深掘りすべきポイントを1つずつ箇条書きで',
        'qa_role': 'あなたはインタビュー記録から質問に答える専門家です。',
        'source_label': 'インタビュー',
    },
    'lecture': {
        'label': '講義',
        'tasks_label': '学習ポイント',
        'summary_role': 'あなたは日本語講義内容を分析する専門家です。',
        'summary_focus': '講義の要点・主要概念・結論',
        'tasks_focus': '学習ポイント・復習すべき項目・重要キーワードを1つずつ箇条書きで',
        'qa_role': 'あなたは講義内容から質問に答える専門家です。',
        'source_label': '講義',
    },
}
ALLOWED_TEMPLATES = set(SUMMARY_TEMPLATES.keys())


def _project_template(conn, project_id: str) -> dict:
    """projects.template から該当テンプレ dict を返す（未設定・不正は meeting）。"""
    row = conn.execute('SELECT template FROM projects WHERE id=?', (project_id,)).fetchone()
    key = (row['template'] if row and row['template'] else 'meeting')
    return SUMMARY_TEMPLATES.get(key, SUMMARY_TEMPLATES['meeting'])


# ── 要約＋残タスク生成 ──────────────────────────────
@app.post('/api/projects/{project_id}/summarize')
async def summarize_project(project_id: str, model: str = 'gemini-2.5-flash',
                            _: str = Depends(require_admin)):
    conn = get_conn()
    merged = conn.execute(
        'SELECT * FROM merged_transcripts WHERE project_id=? ORDER BY id DESC LIMIT 1',
        (project_id,)
    ).fetchone()
    if not merged:
        conn.close()
        raise HTTPException(404, 'No merged transcript. Run /merge first.')

    result = json.loads(merged['result_json']) if merged['result_json'] else []
    if not result:
        conn.close()
        raise HTTPException(400, 'Empty merged transcript')

    # トランスクリプトを単純テキストに整形
    lines = [f'{s.get("ts","")} [{s.get("speaker","")}] {s.get("text","")}' for s in result]
    transcript_text = '\n'.join(lines)

    tpl = _project_template(conn, project_id)
    system_prompt = (
        f"{tpl['summary_role']}"
        '与えられた文字起こしから、要約と項目抽出を行ってください。'
    )
    user_prompt = f"""## {tpl['source_label']}の文字起こし
{transcript_text}

## 出力指示
以下のJSON形式のみで出力してください（説明文・コードフェンス不要）:
{{
  "summary": "200〜400字程度の要約（{tpl['summary_focus']}を含む）",
  "tasks": [
    "{tpl['tasks_focus']}",
    "該当が無い場合は空配列 []"
  ]
}}"""

    logger.info(f'summarize start project={project_id} model={model} template={tpl["label"]}')
    raw = await _call_llm(system_prompt, user_prompt, model, max_tokens=2048, project_id=project_id)

    # JSON抽出（オブジェクト形式）
    fence_match = re.search(r'```(?:json)?\s*(.*?)(?:```|\Z)', raw, re.DOTALL)
    if fence_match:
        raw = fence_match.group(1).strip()
    start = raw.find('{')
    end   = raw.rfind('}')
    if start != -1 and end > start:
        raw = raw[start:end+1]

    try:
        data = json.loads(raw)
    except Exception as e:
        logger.error(f'summarize parse error: {e} raw={raw[:300]}')
        conn.close()
        raise HTTPException(502, f'Summary parse error\nRaw head: {raw[:500]}')

    summary = data.get('summary', '').strip()
    tasks   = data.get('tasks', []) or []

    conn.execute(
        'UPDATE merged_transcripts SET summary=?, tasks_json=? WHERE id=?',
        (summary, json.dumps(tasks, ensure_ascii=False), merged['id'])
    )
    conn.commit()
    conn.close()
    logger.info(f'summarize OK summary_len={len(summary)} tasks={len(tasks)}')
    return {'ok': True, 'summary': summary, 'tasks': tasks}


# ── LLM最終版トランスクリプト：話者名のレコード単位編集 ──────────────
@app.patch('/api/projects/{project_id}/merged/speaker')
async def update_merged_speaker(project_id: str, request: Request,
                                _: str = Depends(require_admin)):
    """LLM最終版トランスクリプトの指定行の話者名を更新する。
    body: {"row_idx": int, "speaker": str}
    """
    body = await request.json()
    row_idx = body.get('row_idx')
    new_speaker = (body.get('speaker') or '').strip()
    if not isinstance(row_idx, int) or row_idx < 0:
        raise HTTPException(400, 'row_idx required (non-negative int)')
    if not new_speaker:
        raise HTTPException(400, 'speaker required (non-empty)')

    conn = get_conn()
    merged = conn.execute(
        'SELECT id, result_json FROM merged_transcripts '
        'WHERE project_id=? ORDER BY id DESC LIMIT 1',
        (project_id,),
    ).fetchone()
    if not merged:
        conn.close()
        raise HTTPException(404, 'merged transcript not found')

    try:
        rows = json.loads(merged['result_json']) if merged['result_json'] else []
    except Exception:
        conn.close()
        raise HTTPException(500, 'result_json parse error')

    if row_idx >= len(rows):
        conn.close()
        raise HTTPException(400, f'row_idx out of range: {row_idx} >= {len(rows)}')

    old_speaker = rows[row_idx].get('speaker', '')
    rows[row_idx]['speaker'] = new_speaker
    txt = rows[row_idx].get('text', '')
    if isinstance(txt, str) and '【要確認:話者】' in txt:
        rows[row_idx]['text'] = txt.replace('【要確認:話者】', '').rstrip()

    conn.execute(
        'UPDATE merged_transcripts SET result_json=? WHERE id=?',
        (json.dumps(rows, ensure_ascii=False), merged['id']),
    )
    conn.commit()
    conn.close()
    logger.info(
        f'merged speaker update: project={project_id} row={row_idx} '
        f'{old_speaker!r} -> {new_speaker!r}'
    )
    return {'ok': True, 'row_idx': row_idx,
            'old_speaker': old_speaker, 'new_speaker': new_speaker}


# ── LLM最終版トランスクリプト：レコードのテキスト更新（内容クリア用） ──────
@app.patch('/api/projects/{project_id}/merged/text')
async def update_merged_text(project_id: str, request: Request,
                             _: str = Depends(require_admin)):
    """LLM最終版トランスクリプトの指定行のテキストを更新する（空文字＝内容クリア）。
    body: {"row_idx": int, "text": str}
    speaker 編集と違い空文字を許可する（重複セルのクリアが目的）。
    行・ts・話者は保持し、Run1/Run2/WS 等の別ソースには影響しない。
    """
    body = await request.json()
    row_idx = body.get('row_idx')
    new_text = body.get('text')
    if not isinstance(row_idx, int) or row_idx < 0:
        raise HTTPException(400, 'row_idx required (non-negative int)')
    if not isinstance(new_text, str):
        raise HTTPException(400, 'text required (str)')

    conn = get_conn()
    merged = conn.execute(
        'SELECT id, result_json FROM merged_transcripts '
        'WHERE project_id=? ORDER BY id DESC LIMIT 1',
        (project_id,),
    ).fetchone()
    if not merged:
        conn.close()
        raise HTTPException(404, 'merged transcript not found')

    try:
        rows = json.loads(merged['result_json']) if merged['result_json'] else []
    except Exception:
        conn.close()
        raise HTTPException(500, 'result_json parse error')

    if row_idx >= len(rows):
        conn.close()
        raise HTTPException(400, f'row_idx out of range: {row_idx} >= {len(rows)}')

    old_text = rows[row_idx].get('text', '')
    rows[row_idx]['text'] = new_text

    conn.execute(
        'UPDATE merged_transcripts SET result_json=? WHERE id=?',
        (json.dumps(rows, ensure_ascii=False), merged['id']),
    )
    conn.commit()
    conn.close()
    logger.info(
        f'merged text update: project={project_id} row={row_idx} '
        f'len {len(old_text)} -> {len(new_text)}'
    )
    return {'ok': True, 'row_idx': row_idx, 'text': new_text}


# ── Excel エクスポート（管理画面のタイムライン4カラムビューをそのまま書出） ──
@app.post('/api/projects/{project_id}/export.xlsx')
async def export_project_xlsx(project_id: str, request: Request,
                              _: str = Depends(require_admin)):
    """管理画面が組み立てた表データをそのままxlsxにする。
    body: { meta:{name,model,summary,tasks,notes},
            headers:[str,str,str,str],
            rows:[{ts, cells:[str,str,str,str]}, ...] }
    """
    from io import BytesIO
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    payload = await request.json()
    meta    = payload.get('meta') or {}
    headers = payload.get('headers') or ['★ LLM最終版', 'Web Speech', 'Run1', 'Run2']
    rows    = payload.get('rows') or []
    summary = (meta.get('summary') or '').strip()
    tasks   = meta.get('tasks') or []
    notes   = meta.get('notes') or []
    name    = meta.get('name') or 'meeting'
    model   = meta.get('model') or ''

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4B0082')
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical='top')

    # === Transcript（タイムライン4カラム比較） ===
    ws = wb.active
    ws.title = 'Transcript'
    full_headers = ['ts'] + list(headers)
    ws.append(full_headers)
    for col_idx in range(1, len(full_headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    for r in rows:
        ts = r.get('ts') or ''
        cells = list(r.get('cells') or [])
        cells = (cells + ['', '', '', ''])[:4]
        ws.append([ts] + cells)
    for i, w in enumerate([8, 50, 50, 50, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.alignment = wrap_top
            cell.border = border
    ws.freeze_panes = 'B2'

    # === Summary シート ===
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 100
    if model:
        ws2.cell(row=1, column=1, value=f'モデル: {model}').font = Font(italic=True, color='666666')
    ws2.cell(row=3, column=1, value=summary or '（要約未生成）').alignment = wrap_top

    # === Tasks シート ===
    ws3 = wb.create_sheet('Tasks')
    ws3.append(['#', '内容'])
    for col_idx in (1, 2):
        c = ws3.cell(row=1, column=col_idx)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = border
    for i, t in enumerate(tasks, 1):
        ws3.append([i, str(t)])
    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 80
    for row in ws3.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.alignment = wrap_top
            cell.border = border

    # === Issues シート（要確認） ===
    ws4 = wb.create_sheet('Issues')
    ws4.append(['ts', '話者', 'note'])
    for col_idx in (1, 2, 3):
        c = ws4.cell(row=1, column=col_idx)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = border
    for n in notes:
        ws4.append([n.get('ts', ''), n.get('speaker', ''), n.get('note', '')])
    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r'[\\/:*?"<>|]', '_', name)[:80]
    ts = datetime.now(timezone(timedelta(hours=9))).strftime('%Y%m%d-%H%M')
    fname = f'{safe_name}_{ts}.xlsx'
    cd = f"attachment; filename=\"export.xlsx\"; filename*=UTF-8''{quote(fname)}"

    return Response(
        content=buf.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': cd}
    )


# ── Q&A（文字起こしへの質問） ─────────────────────
@app.post('/api/projects/{project_id}/ask')
async def ask_project(project_id: str, request: Request,
                      model: str = 'gemini-2.5-flash',
                      _: str = Depends(require_admin)):
    body = await request.json()
    question = (body.get('question') or '').strip()
    if not question:
        raise HTTPException(400, 'question is required')
    if len(question) > 500:
        raise HTTPException(400, 'question too long (max 500 chars)')

    conn = get_conn()
    merged = conn.execute(
        'SELECT * FROM merged_transcripts WHERE project_id=? ORDER BY id DESC LIMIT 1',
        (project_id,)
    ).fetchone()
    if not merged:
        conn.close()
        raise HTTPException(404, 'No merged transcript. Run /merge first.')

    result = json.loads(merged['result_json']) if merged['result_json'] else []
    summary = merged['summary'] or ''
    lines = [f'{s.get("ts","")} [{s.get("speaker","")}] {s.get("text","")}' for s in result]
    transcript_text = '\n'.join(lines) or '（文字起こしなし）'

    tpl = _project_template(conn, project_id)
    system_prompt = (
        f"{tpl['qa_role']}"
        '提供された文字起こしの内容に基づいてのみ回答してください。'
        '推測ではなく、文字起こしに書かれていることだけを根拠にしてください。'
        '不明な場合は「文字起こしには記載がありません」と正直に答えてください。'
    )
    user_prompt = f"""## {tpl['source_label']}の要約
{summary or '（要約未生成）'}

## 文字起こし全文
{transcript_text}

## 質問
{question}

## 回答指示
- 簡潔・自然な日本語で
- 該当する発言があれば「[話者] xxx」と引用する
- 長くても400字以内"""

    logger.info(f'ask start project={project_id} q_len={len(question)} template={tpl["label"]}')
    answer = await _call_llm(system_prompt, user_prompt, model, max_tokens=2048, project_id=project_id)
    logger.info(f'ask OK answer_len={len(answer)}')

    conn.execute(
        'INSERT INTO qa_history(project_id,question,answer,model,created_at) VALUES(?,?,?,?,?)',
        (project_id, question, answer, model, now_iso())
    )
    conn.commit()
    conn.close()
    return {'ok': True, 'answer': answer}


# ── プロジェクト一覧（管理画面） ────────────────────
@app.get('/api/projects')
async def list_projects(_: str = Depends(require_admin)):
    conn = get_conn()
    rows = conn.execute(
        "SELECT p.id, p.name, p.created_at, p.status, "
        "GROUP_CONCAT(pt.name, ',') AS pnames "
        "FROM projects p "
        "LEFT JOIN participants pt ON pt.project_id=p.id "
        "GROUP BY p.id ORDER BY p.created_at DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── プロジェクト詳細（管理画面） ────────────────────
@app.get('/api/projects/{project_id}')
async def get_project(project_id: str, _: str = Depends(require_admin)):
    conn = get_conn()
    proj = conn.execute('SELECT * FROM projects WHERE id=?', (project_id,)).fetchone()
    if not proj:
        conn.close()
        raise HTTPException(404, 'Project not found')

    participants = [dict(r) for r in conn.execute(
        'SELECT idx,name FROM participants WHERE project_id=? ORDER BY idx', (project_id,)
    ).fetchall()]

    recordings_raw = [dict(r) for r in conn.execute(
        'SELECT id,seq,mime,duration,created_at,audio_path,zoom_participant_name '
        'FROM recordings WHERE project_id=? ORDER BY seq',
        (project_id,)
    ).fetchall()]
    recordings = []
    for r in recordings_raw:
        # 表示名（順序入替後も識別可能な安定値）
        basename = os.path.basename(r.get('audio_path') or '')
        # 先頭の "01_" のような連番prefix除去（追加アップロードで付与される）
        stripped = re.sub(r'^\d{1,3}_', '', basename)
        # モバイル録音は "{project_id}_{seq}.{ext}" 形式 → 友好名なしと判定
        is_mobile_path = bool(
            basename.startswith(f"{project_id}_") or
            re.match(r'^[a-f0-9-]{20,}_\d+\.[a-z0-9]+$', basename, re.IGNORECASE)
        )
        # 拡張子も削る
        name_no_ext = re.sub(r'\.[a-z0-9]+$', '', stripped, flags=re.IGNORECASE)
        if r.get('zoom_participant_name'):
            display_name = r['zoom_participant_name']
        elif is_mobile_path or not name_no_ext:
            display_name = f"作成 {(r.get('created_at') or '')[5:16].replace('T',' ')}"
        else:
            display_name = name_no_ext
        # UI で確実に表示揃えしたいので 24文字で切り詰め（…付き）
        if len(display_name) > 24:
            display_name = display_name[:23] + '…'
        r['display_name'] = display_name
        # audio_path はパス漏洩防止で除外
        r.pop('audio_path', None)
        recordings.append(r)

    ref_segs = [dict(r) for r in conn.execute(
        'SELECT seq,text,speaker_idx,ts,highlight FROM ref_segments WHERE project_id=? ORDER BY seq',
        (project_id,)
    ).fetchall()]

    txs_raw = conn.execute(
        'SELECT * FROM ai_transcripts WHERE project_id=? ORDER BY run_number,id', (project_id,)
    ).fetchall()
    transcripts = []
    for tx in txs_raw:
        t = dict(tx)
        t['utterances']  = json.loads(t['utterances_json'])  if t['utterances_json']  else []
        t['speaker_map'] = json.loads(t['speaker_map_json']) if t['speaker_map_json'] else {}
        del t['utterances_json'], t['speaker_map_json']
        transcripts.append(t)

    # LLM突合結果（最新1件）
    merged_raw = conn.execute(
        'SELECT * FROM merged_transcripts WHERE project_id=? ORDER BY id DESC LIMIT 1',
        (project_id,)
    ).fetchone()
    merged = None
    if merged_raw:
        merged = dict(merged_raw)
        merged['result'] = json.loads(merged['result_json']) if merged['result_json'] else []
        merged['notes']  = json.loads(merged['notes_json'])  if merged['notes_json']  else []
        merged['tasks']  = json.loads(merged['tasks_json'])  if merged.get('tasks_json') else []
        del merged['result_json'], merged['notes_json']
        if 'tasks_json' in merged: del merged['tasks_json']

    # Q&A履歴
    qa = [dict(r) for r in conn.execute(
        'SELECT id, question, answer, model, created_at FROM qa_history '
        'WHERE project_id=? ORDER BY id ASC', (project_id,)
    ).fetchall()]

    conn.close()
    return {
        **dict(proj),
        'participants': participants,
        'recordings':   recordings,
        'ref_segments': ref_segs,
        'transcripts':  transcripts,
        'merged':       merged,
        'qa':           qa,
    }


# ── MTG名変更 ────────────────────────────────────
@app.patch('/api/projects/{project_id}')
async def update_project(project_id: str, request: Request, _: str = Depends(require_admin)):
    body = await request.json()
    conn = get_conn()
    proj = conn.execute('SELECT id FROM projects WHERE id=?', (project_id,)).fetchone()
    if not proj:
        conn.close()
        raise HTTPException(404, 'Project not found')
    if 'name' in body:
        conn.execute('UPDATE projects SET name=? WHERE id=?', (body['name'].strip(), project_id))
        conn.commit()
        logger.info(f'Project renamed: {project_id} → {body["name"]}')
    if 'template' in body:
        tmpl = (body.get('template') or '').strip()
        if tmpl not in ALLOWED_TEMPLATES:
            conn.close()
            raise HTTPException(400, f'template must be one of: {sorted(ALLOWED_TEMPLATES)}')
        conn.execute('UPDATE projects SET template=? WHERE id=?', (tmpl, project_id))
        conn.commit()
        logger.info(f'Project template set: {project_id} → {tmpl}')
    conn.close()
    return {'ok': True}


# ── MTG削除 ──────────────────────────────────────
@app.delete('/api/projects/{project_id}')
async def delete_project(project_id: str, _: str = Depends(require_admin)):
    conn = get_conn()
    proj = conn.execute('SELECT id FROM projects WHERE id=?', (project_id,)).fetchone()
    if not proj:
        conn.close()
        raise HTTPException(404, 'Project not found')
    # 音声ファイルをディスクから削除
    recs = conn.execute('SELECT audio_path FROM recordings WHERE project_id=?', (project_id,)).fetchall()
    for rec in recs:
        try:
            os.remove(rec['audio_path'])
        except Exception:
            pass
    # DB レコードを全削除（外部キー順）
    for table in ('qa_history', 'merged_transcripts', 'ai_transcripts', 'ref_segments',
                  'speaker_events', 'participants', 'recordings'):
        conn.execute(f'DELETE FROM {table} WHERE project_id=?', (project_id,))
    conn.execute('DELETE FROM projects WHERE id=?', (project_id,))
    conn.commit()
    conn.close()
    logger.info(f'Project deleted: {project_id}')
    return {'ok': True}


# ── 録音音声 追加アップロード ────────────────────
@app.post('/api/projects/{project_id}/recordings')
async def add_recordings(project_id: str, request: Request,
                          _: str = Depends(require_admin)):
    """既存プロジェクトに音声ファイルを追加する。

    multipart:
      - files: 音声ファイル（複数可）
      - meta: JSON 文字列 {speakers:[{filename, name}]} 任意
    """
    form = await request.form()
    files = form.getlist('files')
    if not files:
        raise HTTPException(400, 'no files uploaded')
    try:
        meta = json.loads(form.get('meta', '{}'))
    except Exception:
        meta = {}
    name_map = {p['filename']: p['name']
                for p in meta.get('speakers', []) if 'filename' in p}

    conn = get_conn()
    proj = conn.execute('SELECT id FROM projects WHERE id=?', (project_id,)).fetchone()
    if not proj:
        conn.close()
        raise HTTPException(404, 'project not found')

    max_seq_row = conn.execute(
        'SELECT COALESCE(MAX(seq),-1) AS m FROM recordings WHERE project_id=?',
        (project_id,)
    ).fetchone()
    next_seq = int(max_seq_row['m']) + 1
    max_part_row = conn.execute(
        'SELECT COALESCE(MAX(idx),0) AS m FROM participants WHERE project_id=?',
        (project_id,)
    ).fetchone()
    next_part_idx = int(max_part_row['m']) + 1

    proj_dir = f'/var/jizo/audio/{project_id}'
    os.makedirs(proj_dir, exist_ok=True)

    added = []
    for f in files:
        fname = getattr(f, 'filename', f'audio_{next_seq}.m4a')
        ext = os.path.splitext(fname)[1].lstrip('.').lower() or 'm4a'
        dst = os.path.join(proj_dir, f'{next_seq:02d}_{fname}')
        content = await f.read()
        with open(dst, 'wb') as out:
            out.write(content)
        mime = getattr(f, 'content_type', None) or f'audio/{ext}'
        speaker = name_map.get(fname) or os.path.splitext(fname)[0]

        cur = conn.execute(
            'INSERT INTO recordings(project_id,seq,audio_path,mime,duration,'
            'created_at,zoom_participant_name) VALUES(?,?,?,?,?,?,?)',
            (project_id, next_seq, dst, mime, '', now_iso(), speaker),
        )
        conn.execute(
            'INSERT INTO participants(project_id,idx,name) VALUES(?,?,?)',
            (project_id, next_part_idx, speaker),
        )
        added.append({'recording_id': cur.lastrowid, 'seq': next_seq,
                      'speaker': speaker, 'filename': fname})
        next_seq += 1
        next_part_idx += 1

    conn.execute(
        'UPDATE projects SET recordings_updated_at=? WHERE id=?',
        (now_iso(), project_id),
    )
    conn.commit()
    conn.close()
    logger.info(f'recordings added: project={project_id} count={len(added)}')
    return {'ok': True, 'added': added}


# ── 録音音声 順序入替 ────────────────────────────
@app.patch('/api/projects/{project_id}/recordings/order')
async def reorder_recordings(project_id: str, request: Request,
                              _: str = Depends(require_admin)):
    """recording_id 配列の順で seq を 0..N-1 に再割り当て。

    body: {"order": [recording_id, ...]}
    """
    body = await request.json()
    order = body.get('order', [])
    if not isinstance(order, list) or not order:
        raise HTTPException(400, 'order must be non-empty list of recording_id')

    conn = get_conn()
    rows = conn.execute(
        'SELECT id FROM recordings WHERE project_id=?', (project_id,)
    ).fetchall()
    existing_ids = {r['id'] for r in rows}
    given_ids = set(order)
    if existing_ids != given_ids:
        conn.close()
        raise HTTPException(
            400,
            f'order must contain exactly all recordings of this project. '
            f'expected={sorted(existing_ids)} given={sorted(given_ids)}',
        )

    # 衝突回避：いったん全部マイナスに退避
    conn.execute(
        'UPDATE recordings SET seq = -1000 - seq WHERE project_id=?',
        (project_id,),
    )
    for new_seq, rid in enumerate(order):
        conn.execute(
            'UPDATE recordings SET seq=? WHERE id=? AND project_id=?',
            (new_seq, rid, project_id),
        )
    conn.execute(
        'UPDATE projects SET recordings_updated_at=? WHERE id=?',
        (now_iso(), project_id),
    )
    conn.commit()
    conn.close()
    logger.info(f'recordings reordered: project={project_id} count={len(order)}')
    return {'ok': True, 'order': order}


# ── 音声ファイル配信（管理画面） ────────────────────
@app.get('/api/audio/{project_id}/{seq}')
async def get_audio(project_id: str, seq: int, _: str = Depends(require_admin)):
    conn = get_conn()
    rec = conn.execute(
        'SELECT * FROM recordings WHERE project_id=? AND seq=?', (project_id, seq)
    ).fetchone()
    conn.close()
    if not rec or not os.path.exists(rec['audio_path']):
        raise HTTPException(404, 'Audio not found')
    return FileResponse(rec['audio_path'], media_type=rec['mime'])


# ── Zoom 連携：録画ファイル DL ワーカー ─────────────
async def _zoom_download_worker(pending_id: int, payload: dict, dl_token: str):
    """recording_files から participant_audio_files を抽出して個別 DL する。

    話者別音声ファイル（participant_audio）が無い場合は audio_only にフォールバック
    （話者分離なし＝既存モバイル録音と同等のフロー）。
    """
    uuid_     = payload.get('uuid', '')
    safe_uuid = re.sub(r'[^A-Za-z0-9_-]', '_', uuid_)
    out_dir   = os.path.join(ZOOM_AUDIO_DIR, safe_uuid)
    os.makedirs(out_dir, exist_ok=True)

    files = payload.get('recording_files', [])

    # 話者別音声を優先（recording_type=audio_interpretation 等の特殊型も participant_email がある場合は含める）
    p_audio = [f for f in files
               if f.get('recording_type') == 'participant_audio'
               or f.get('participant_email')]
    targets = p_audio if p_audio else [
        f for f in files if f.get('recording_type') == 'audio_only'
    ]

    if not p_audio:
        logger.warning(
            f'zoom dl {uuid_}: no participant_audio (fallback to mixed audio)'
        )

    participants: list[dict] = []
    try:
        for f in targets:
            url = f.get('download_url')
            if not url:
                continue
            raw_name = (
                f.get('participant_email')
                or f.get('file_name')
                or f.get('id', 'unknown')
            )
            # ファイル名安全化（英数・日本語・記号一部のみ許可）
            safe_name = re.sub(r'[^\w\-.@ぁ-んァ-ヴー一-龥]', '_', raw_name)
            ext = (f.get('file_extension') or 'm4a').lower()
            dest = os.path.join(out_dir, f'{safe_name}.{ext}')

            await zoom_client.download_file(url, dest, token=dl_token)
            participants.append({
                'name': safe_name,
                'audio_path': dest,
                'file_size': int(f.get('file_size', 0)),
            })

        conn = get_conn()
        conn.execute("""
            UPDATE zoom_pending
               SET participants_json=?, downloaded_at=?, status='ready'
             WHERE id=?
        """, (
            json.dumps(participants, ensure_ascii=False),
            datetime.now(timezone.utc).isoformat(),
            pending_id,
        ))
        conn.commit()
        conn.close()
        logger.info(f'zoom dl {uuid_}: {len(participants)} files done')

    except Exception as e:
        logger.exception(f'zoom dl {uuid_} failed: {e}')
        conn = get_conn()
        conn.execute(
            "UPDATE zoom_pending SET status='error', error_message=? WHERE id=?",
            (str(e), pending_id),
        )
        conn.commit()
        conn.close()


# ── Zoom 連携：Webhook 受信（通知のみ・自動DLしない） ─────
@app.post('/api/zoom/webhook')
async def zoom_webhook(request: Request):
    """Zoom Webhook 受信エンドポイント（通知のみ）。

    - `endpoint.url_validation`: 初期検証チャレンジに即時応答
    - `recording.completed`: 署名検証 → zoom_pending(status='notified') を記録するのみ
                            （自動DLしない。実際の取り込みは管理画面から明示実行）
    - その他イベント: 無視
    """
    body = await request.body()
    ts   = request.headers.get('x-zm-request-timestamp', '')
    sig  = request.headers.get('x-zm-signature', '')

    try:
        data = json.loads(body)
    except Exception:
        raise HTTPException(400, 'invalid json')

    # 1) URL Validation Challenge（署名検証より先に処理）
    if data.get('event') == 'endpoint.url_validation':
        plain = data.get('payload', {}).get('plainToken', '')
        return build_url_validation_response(plain, ZOOM_WEBHOOK_SECRET)

    # 2) 通常イベントの署名検証
    if not verify_signature(body, ts, sig, ZOOM_WEBHOOK_SECRET):
        logger.warning(f'zoom webhook: invalid signature ts={ts}')
        raise HTTPException(401, 'invalid signature')

    # 3) recording.completed 以外は無視
    if data.get('event') != 'recording.completed':
        logger.info(f"zoom webhook: ignored event={data.get('event')}")
        return {'status': 'ignored'}

    payload = data.get('payload', {}).get('object', {})
    uuid_ = payload.get('uuid', '')

    # 通知のみ記録（DLしない。download_token は時間切れになる前提で保存しない）
    conn = get_conn()
    try:
        conn.execute("""
            INSERT INTO zoom_pending
              (zoom_meeting_id, zoom_uuid, topic, host_email, start_time,
               duration, status, created_at)
            VALUES (?,?,?,?,?,?,'notified',?)
        """, (
            str(payload.get('id', '')), uuid_,
            payload.get('topic', ''), payload.get('host_email', ''),
            payload.get('start_time', ''), int(payload.get('duration', 0)),
            datetime.now(timezone.utc).isoformat(),
        ))
        conn.commit()
        logger.info(f'zoom webhook: notified uuid={uuid_} topic={payload.get("topic","")}')
    except Exception as e:
        # UNIQUE 違反（同じ MTG の重複 Webhook）は握りつぶす
        logger.info(f'zoom webhook: duplicate uuid {uuid_} ({e})')
    finally:
        conn.close()
    return {'status': 'notified'}


# ── Zoom 連携：取り込み待ち一覧 ─────────────────────
@app.get('/api/zoom/pending')
async def zoom_pending_list(_: str = Depends(require_admin)):
    """取り込み待ち（pending_download / ready / error）の Zoom MTG を返す"""
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, zoom_meeting_id, zoom_uuid, topic, host_email,
               start_time, duration, participants_json, status,
               error_message, created_at, imported_project_id
          FROM zoom_pending
         WHERE status IN ('pending_download','ready','error')
         ORDER BY created_at DESC
    """).fetchall()
    conn.close()

    out = []
    for r in rows:
        d = dict(r)
        try:
            d['participants'] = json.loads(d.pop('participants_json') or '[]')
        except Exception:
            d['participants'] = []
        out.append(d)
    return {'pending': out}


# ── Zoom 連携：Zoom Cloud 録画一覧（API + Webhook通知マージ） ─────
@app.get('/api/zoom/cloud-recordings')
async def zoom_cloud_recordings(days: int = 30, _: str = Depends(require_admin)):
    """Zoom Cloud の録画一覧と zoom_pending(notified等) をマージして返す。

    - Zoom API 由来: 取り込み可能（recording_files が取れる）
    - Webhook 通知のみ（Zoom API に無い）: 削除済の可能性。表示はするが取り込み不可
    マージキー: zoom_uuid
    """
    if zoom_client is None:
        raise HTTPException(503, 'zoom integration not configured')

    # Zoom API（自アカウントの録画のみ）
    tok = await zoom_client.get_access_token()
    from_date = (datetime.now(timezone.utc) - timedelta(days=max(1, min(days, 365)))).strftime('%Y-%m-%d')
    zoom_meetings: list[dict] = []
    try:
        async with httpx.AsyncClient(timeout=30) as cli:
            r = await cli.get(
                'https://api.zoom.us/v2/users/me/recordings',
                headers={'Authorization': f'Bearer {tok}'},
                params={'page_size': 30, 'from': from_date},
            )
            r.raise_for_status()
            zoom_meetings = r.json().get('meetings', [])
    except Exception as e:
        logger.exception(f'zoom cloud-recordings: API fetch failed: {e}')
        raise HTTPException(502, f'Zoom API error: {e}')

    # zoom_pending（webhook 通知ログ・取り込み済を含む）
    conn = get_conn()
    rows = conn.execute("""
        SELECT zoom_uuid, status, imported_project_id, error_message,
               created_at, topic, start_time, host_email
          FROM zoom_pending
    """).fetchall()
    conn.close()
    pending_by_uuid = {r['zoom_uuid']: dict(r) for r in rows}

    out = []
    seen_uuids = set()
    for m in zoom_meetings:
        u = m.get('uuid', '')
        seen_uuids.add(u)
        local = pending_by_uuid.get(u, {})
        files = m.get('recording_files', [])
        p_audio = [f for f in files
                   if f.get('recording_type') == 'participant_audio'
                   or f.get('participant_email')]
        out.append({
            'source': 'zoom_api',
            'uuid': u,
            'meeting_id': m.get('id'),
            'topic': m.get('topic'),
            'host_email': m.get('host_email'),
            'start_time': m.get('start_time'),
            'duration_min': m.get('duration'),
            'file_count': len(files),
            'participant_audio_count': len(p_audio),
            'has_participant_audio': len(p_audio) > 0,
            'status': local.get('status') or 'available',
            'imported_project_id': local.get('imported_project_id'),
            'webhook_notified_at': local.get('created_at'),
        })

    # Zoom API に無いが webhook 通知だけある（= Zoom側で削除済の可能性）
    for u, local in pending_by_uuid.items():
        if u in seen_uuids:
            continue
        out.append({
            'source': 'webhook_only',
            'uuid': u,
            'meeting_id': None,
            'topic': local.get('topic'),
            'host_email': local.get('host_email'),
            'start_time': local.get('start_time'),
            'duration_min': None,
            'file_count': 0,
            'participant_audio_count': 0,
            'has_participant_audio': False,
            'status': local.get('status'),
            'imported_project_id': local.get('imported_project_id'),
            'webhook_notified_at': local.get('created_at'),
            'note': 'Zoom側に録画が見つかりません（削除済の可能性）',
        })

    out.sort(key=lambda x: (x.get('start_time') or x.get('webhook_notified_at') or ''), reverse=True)
    return {'recordings': out, 'from_date': from_date, 'days': days}


# ── Zoom 連携：zoom_pending → projects 変換（内部関数） ─────
def _zoom_pending_to_project(pending_id: int) -> dict:
    """zoom_pending(status='ready') → projects/recordings/participants 化する。

    STT 実行は行わない。呼び出し元エンドポイントの責務。
    """
    conn = get_conn()
    row = conn.execute(
        "SELECT * FROM zoom_pending WHERE id=? AND status='ready'",
        (pending_id,),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, 'pending not found or not ready')

    topic        = row['topic'] or 'Zoom MTG'
    start_time   = row['start_time'] or datetime.now(timezone.utc).isoformat()
    project_id   = uuid.uuid4().hex
    project_name = f'{topic} ({start_time[:16]})'
    proj_dir     = f'/var/jizo/audio/{project_id}'
    os.makedirs(proj_dir, exist_ok=True)

    conn.execute(
        '''INSERT INTO projects(id, name, created_at, status, source)
           VALUES(?, ?, ?, ?, 'zoom')''',
        (project_id, project_name,
         datetime.now(timezone.utc).isoformat(), 'uploaded'),
    )

    participants = json.loads(row['participants_json'] or '[]')
    rec_ids: list[int] = []
    for idx, p in enumerate(participants, start=1):
        src = p['audio_path']
        ext = os.path.splitext(src)[1] or '.m4a'
        dst = os.path.join(proj_dir, f'{idx:02d}_{os.path.basename(src)}')
        try:
            os.rename(src, dst)
        except OSError:
            import shutil
            shutil.copy2(src, dst)

        mime = 'audio/m4a' if ext.lower() in ('.m4a', '.mp4') else 'audio/wav'
        cur = conn.execute(
            '''INSERT INTO recordings
               (project_id, seq, audio_path, mime, duration, created_at,
                zoom_participant_name)
               VALUES(?,?,?,?,?,?,?)''',
            (project_id, idx, dst, mime, '',
             datetime.now(timezone.utc).isoformat(),
             p.get('name', '')),
        )
        rec_ids.append(cur.lastrowid)
        conn.execute(
            'INSERT INTO participants(project_id, idx, name) VALUES(?,?,?)',
            (project_id, idx, p.get('name', '')),
        )

    conn.execute(
        '''UPDATE zoom_pending
              SET status='imported', imported_project_id=?
            WHERE id=?''',
        (project_id, pending_id),
    )
    conn.commit()
    conn.close()

    logger.info(f'zoom import: project={project_id} recordings={len(rec_ids)}')
    return {
        'project_id': project_id,
        'recordings': len(rec_ids),
        'name': project_name,
    }


# ── Zoom 連携：取り込み実行（旧フロー・後方互換） ─────
@app.post('/api/zoom/pending/{pending_id}/import')
async def zoom_import(pending_id: int, _: str = Depends(require_admin)):
    """旧フロー（Webhook自動DL→ready→import）の互換エンドポイント。

    新フローは POST /api/zoom/cloud-recordings/{uuid}/import を使用。
    """
    return _zoom_pending_to_project(pending_id)


# ── Zoom 連携：UUID から録画 meeting を取得（一覧 API 経由） ─────
async def _zoom_fetch_recording_by_uuid(uuid: str, months_back: int = 6):
    """指定 UUID の録画 meeting dict を GET /users/me/recordings から探して返す。

    S2S OAuth アプリでは GET /meetings/{uuid}/recordings が granular scope
    `cloud_recording:read:list_recording_files(:admin)` を要求するが、この scope は
    S2S アプリに付与できない（Zoom 仕様・code 4711）。そのため取り込み済 scope
    `cloud_recording:read:list_user_recordings:admin` で叩ける一覧 API を、
    月単位ウィンドウ（Zoom の from/to は最大1ヶ月）で遡って検索する。
    recording_files（download_url 含む）をそのまま返す。見つからなければ None。
    """
    tok = await zoom_client.get_access_token()
    today = datetime.now(timezone.utc).date()
    async with httpx.AsyncClient(timeout=30) as cli:
        for i in range(max(1, months_back)):
            to_d = today - timedelta(days=30 * i)
            from_d = to_d - timedelta(days=31)
            page_token = ''
            while True:
                params = {
                    'page_size': 300,
                    'from': from_d.isoformat(),
                    'to': to_d.isoformat(),
                }
                if page_token:
                    params['next_page_token'] = page_token
                r = await cli.get(
                    'https://api.zoom.us/v2/users/me/recordings',
                    headers={'Authorization': f'Bearer {tok}'},
                    params=params,
                )
                if r.status_code != 200:
                    logger.warning(
                        f'zoom find-recording: list HTTP {r.status_code} '
                        f'window {from_d}..{to_d}: {r.text[:160]}'
                    )
                    break
                data = r.json()
                for m in data.get('meetings', []):
                    if m.get('uuid') == uuid:
                        return m
                page_token = data.get('next_page_token') or ''
                if not page_token:
                    break
    return None


# ── Zoom 連携：Cloud 録画から取り込み（新フロー・1ステップ） ─────
@app.post('/api/zoom/cloud-recordings/{uuid:path}/import')
async def zoom_cloud_import(uuid: str, _: str = Depends(require_admin)):
    """指定 UUID の Zoom 録画を Zoom Cloud から DL → projects 化（1ステップ）。

    OAuth Bearer で DL するため Webhook の download_token は不要。
    既に取り込み済の場合は 409 を返さず {ok:false, project_id} を返す。
    """
    if zoom_client is None:
        raise HTTPException(503, 'zoom integration not configured')

    # 1) 既存チェック（imported なら即返却）
    conn = get_conn()
    existing = conn.execute(
        "SELECT id, status, imported_project_id FROM zoom_pending WHERE zoom_uuid=?",
        (uuid,),
    ).fetchone()
    if existing and existing['status'] == 'imported':
        conn.close()
        return {
            'ok': False,
            'message': 'already imported',
            'project_id': existing['imported_project_id'],
        }
    conn.close()

    # 2) Zoom 録画情報を取得（一覧 API 経由）
    #    GET /meetings/{uuid}/recordings は S2S アプリに付与できない scope を要求するため使わず、
    #    GET /users/me/recordings（list_user_recordings:admin）から UUID 一致で探す。
    try:
        meeting = await _zoom_fetch_recording_by_uuid(uuid)
    except Exception as e:
        logger.exception(f'zoom cloud-import: Zoom API fetch failed uuid={uuid}: {e}')
        raise HTTPException(502, f'Zoom API error: {e}')
    if meeting is None:
        raise HTTPException(
            404,
            'recording not found in Zoom user recordings '
            '(削除済 / 6ヶ月より前 / このアカウント主催でない 可能性)',
        )

    payload = {
        'uuid': uuid,
        'id': meeting.get('id'),
        'topic': meeting.get('topic', ''),
        'host_email': meeting.get('host_email', ''),
        'start_time': meeting.get('start_time', ''),
        'duration': int(meeting.get('duration', 0) or 0),
        'recording_files': meeting.get('recording_files', []),
    }

    # 3) zoom_pending を upsert（pending_download 状態にしてDLへ）
    conn = get_conn()
    existing = conn.execute(
        "SELECT id FROM zoom_pending WHERE zoom_uuid=?",
        (uuid,),
    ).fetchone()
    if existing:
        pending_id = existing['id']
        conn.execute(
            "UPDATE zoom_pending SET status='pending_download', error_message=NULL WHERE id=?",
            (pending_id,),
        )
    else:
        cur = conn.execute("""
            INSERT INTO zoom_pending
              (zoom_meeting_id, zoom_uuid, topic, host_email, start_time,
               duration, status, created_at)
            VALUES (?,?,?,?,?,?,'pending_download',?)
        """, (
            str(payload['id']), uuid, payload['topic'],
            payload['host_email'], payload['start_time'],
            payload['duration'],
            datetime.now(timezone.utc).isoformat(),
        ))
        pending_id = cur.lastrowid
    conn.commit()
    conn.close()

    # 4) DL（OAuth Bearer 経路。dl_token=None で worker が OAuth に切り替え）
    await _zoom_download_worker(pending_id, payload, dl_token=None)

    # 5) DL 成否確認
    conn = get_conn()
    row = conn.execute(
        "SELECT status, error_message FROM zoom_pending WHERE id=?",
        (pending_id,),
    ).fetchone()
    conn.close()
    if row['status'] != 'ready':
        raise HTTPException(500, f"download failed: {row['error_message']}")

    # 6) projects 化
    result = _zoom_pending_to_project(pending_id)
    return {'ok': True, **result}


# ── Zoom 連携：取り込み待ち破棄 ─────────────────────
@app.delete('/api/zoom/pending/{pending_id}')
async def zoom_pending_delete(pending_id: int,
                              _: str = Depends(require_admin)):
    """取り込みせずに破棄。ダウンロード済み音声ファイルも削除する"""
    conn = get_conn()
    row = conn.execute(
        'SELECT zoom_uuid, participants_json FROM zoom_pending WHERE id=?',
        (pending_id,),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, 'not found')

    try:
        for p in json.loads(row['participants_json'] or '[]'):
            try:
                os.remove(p['audio_path'])
            except OSError:
                pass
        safe_uuid = re.sub(r'[^A-Za-z0-9_-]', '_', row['zoom_uuid'])
        d = os.path.join(ZOOM_AUDIO_DIR, safe_uuid)
        if os.path.isdir(d) and not os.listdir(d):
            os.rmdir(d)
    finally:
        conn.execute('DELETE FROM zoom_pending WHERE id=?', (pending_id,))
        conn.commit()
        conn.close()

    logger.info(f'zoom pending deleted: id={pending_id}')
    return {'deleted': pending_id}


# ── Zoom 連携：手動アップロード（他組織主催 MTG 用） ──
@app.post('/api/zoom/upload')
async def zoom_upload(request: Request,
                      _: str = Depends(require_admin)):
    """multipart/form-data:
       - files[]   : 音声ファイル（.m4a / .mp4 / .wav）
       - meta      : JSON 文字列 {topic, participants:[{filename, name}]}
    """
    form = await request.form()
    try:
        meta = json.loads(form.get('meta', '{}'))
    except Exception:
        raise HTTPException(400, 'invalid meta json')
    files = form.getlist('files')
    if not files:
        raise HTTPException(400, 'no files uploaded')

    project_id = uuid.uuid4().hex
    proj_dir   = f'/var/jizo/audio/{project_id}'
    os.makedirs(proj_dir, exist_ok=True)

    # ファイル名 → 表示話者名のマッピング
    name_map = {p['filename']: p['name']
                for p in meta.get('participants', []) if 'filename' in p}

    topic = meta.get('topic', 'Zoom upload')

    conn = get_conn()
    conn.execute(
        '''INSERT INTO projects(id, name, created_at, status, source)
           VALUES(?, ?, ?, ?, 'upload')''',
        (project_id, topic,
         datetime.now(timezone.utc).isoformat(), 'uploaded'),
    )

    for idx, f in enumerate(files, start=1):
        fname = getattr(f, 'filename', f'audio{idx}.m4a')
        dst   = os.path.join(proj_dir, f'{idx:02d}_{fname}')
        content = await f.read()
        with open(dst, 'wb') as out:
            out.write(content)

        speaker = name_map.get(fname) or os.path.splitext(fname)[0]
        mime = getattr(f, 'content_type', None) or 'audio/m4a'

        conn.execute(
            '''INSERT INTO recordings
               (project_id, seq, audio_path, mime, duration, created_at,
                zoom_participant_name)
               VALUES(?,?,?,?,?,?,?)''',
            (project_id, idx, dst, mime, '',
             datetime.now(timezone.utc).isoformat(), speaker),
        )
        conn.execute(
            'INSERT INTO participants(project_id, idx, name) VALUES(?,?,?)',
            (project_id, idx, speaker),
        )

    conn.commit()
    conn.close()

    logger.info(f'zoom upload: project={project_id} files={len(files)}')
    return {'project_id': project_id, 'recordings': len(files), 'name': topic}
